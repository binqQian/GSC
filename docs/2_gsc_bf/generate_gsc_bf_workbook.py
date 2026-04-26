from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
REPO_ROOT = ROOT.parents[1]
BASE_DATA = REPO_ROOT / "docs" / "1_0423_gsc_binary_format" / "generate_gsc_binary_docs.py"
OUT_XLSX = ROOT / "GSC_VCF_gVCF二进制结构表-260424.xlsx"


def load_base_data():
    spec = importlib.util.spec_from_file_location("gsc_binary_docs_data", BASE_DATA)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {BASE_DATA}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


BASE = load_base_data()

COMMON_COLUMNS = [
    "一级区段",
    "层级路径",
    "顺序",
    "字段/条目",
    "类型/编码",
    "必选/条件",
    "版本/适用条件",
    "说明",
    "备注/兼容",
    "源码依据",
]

STRUCTURE_HEADER_ROW = [
    "Primary Structure",
    None,
    None,
    "Secondary Structure",
    None,
    None,
    "Tertiary Structure",
    None,
    None,
    "Quaternary Structure",
    None,
    None,
]

STRUCTURE_LABEL_ROW = [
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
]

STRUCTURE_WIDTHS = [24, 10, 26, 28, 10, 30, 32, 10, 34, 34, 10, 42]


def srow(*vals: str | None) -> tuple[str | None, ...]:
    padded = list(vals[:12])
    padded.extend([None] * (12 - len(padded)))
    return tuple(padded)

INFO_ROWS = [
    ("文档范围", "覆盖当前仓库真实写出的两类二进制结构：多样本 VCF/BCF -> `.gsc`，单样本 `gsc gvcf` -> native `GVCF`。"),
    ("表格组织", "按用户要求把 VCF `.gsc` 与 gVCF native 二进制结构分别放到 `VCF_gsc二进制结构` 和 `gVCF_native二进制结构` 两个 sheet。"),
    ("参考形式", "参考 `docs/0_avsStructure/基因比对压缩封装格式规范.md` 与 `基因比对压缩封装格式-241227.xlsx` 的结构化表格表达，但字段顺序以源码 writer/reader 为准。"),
    ("VCF 源码依据", "`src/compressor.cpp::writeCompressFlie()`、`src/decompression_reader.cpp::OpenReading()/readFixedFields()`、`src/file_handle.cpp`、`src/variant.h`。"),
    ("gVCF 源码依据", "`src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/CompressAndWriteBlock()/WriteFileFooter()`、`src/gvcf/gvcf_block.cpp`、`src/gvcf/gvcf_field_compress.cpp`、`src/gvcf/gvcf_encoding.cpp`。"),
    ("当前版本", "多样本 fixed-fields chunk 当前 writer 为 `GSCF V2`；native gVCF 当前 writer 为 `GVCF_FILE_VERSION=4`。"),
    ("ABI 注意", "多样本 `.gsc` 主 archive 直接写 `bool`、`size_t`，part2 footer 末尾 `footer_size` 固定写 8 字节，默认依赖 64-bit little-endian ABI。"),
]

VCF_AUX_ROWS = [
    (
        "6_AUXILIARY_ENCODINGS",
        "fixed/meta/part2_params",
        "-",
        "append_str()/read_str()",
        "raw string bytes + NUL",
        "as used",
        "current writer",
        "用于 header、sample names、fixed fields 字符串列、part2 key name 等字符串 token。",
        "不是 length-prefixed string。",
        "src/utils.cpp; src/compressor.cpp::compress_meta(); src/block_processing.cpp::addFixedFieldsBlock()",
    ),
    (
        "6_AUXILIARY_ENCODINGS",
        "fixed/part2_params",
        "-",
        "append()/read() numeric token",
        "decimal ASCII bytes + NUL",
        "as used",
        "current writer",
        "用于 POS delta、actual_variants、key 表、backend id、fmt dict metadata 等数值 token。",
        "不能按定长整数直接解析这些 payload 解压后的内部内容。",
        "src/utils.h; src/block_processing.cpp::addFixedFieldsBlock(); src/compressor.cpp::CompressProcess()",
    ),
    (
        "6_AUXILIARY_ENCODINGS",
        "GT/permutation",
        "-",
        "vint_code::EncodeArray()",
        "variable-length integer bytes",
        "as used",
        "current writer",
        "用于 GT sparse row 的 delta list，以及 row/column block permutation payload。",
        "GT sparse row 中的 `0` 是行终止符，合法 delta 必须大于 0。",
        "src/vint_code.cpp; src/block_processing.cpp; src/decompressor.cpp::decoded_vector_row()",
    ),
]

GVCF_FIELD_ROWS = [
    (
        "3_COMPRESSED_FIELD_BACKEND_WRAPPER",
        "CompressedField/data",
        "-",
        "backend_flag",
        "uint8_t + payload",
        "when data_size > 0",
        "current writer",
        "`CompressedField.data` 内部第 1 字节表示是否再经过通用后端压缩：`0`=后接原始字段编码 payload，`1`=后接 backend-compressed payload。",
        "空字段 `data_size=0` 时没有这个 flag。",
        "src/gvcf/gvcf_field_compress.cpp::FieldCompressor::ApplyBackendCompression(); src/gvcf/gvcf_field_decompress.cpp::ApplyBackendDecompression()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/chrom",
        "6",
        "chrom",
        "CompressedField(method=RLE)",
        "Y",
        "block field",
        "CHROM 字符串数组先做 `RLE(string)`，再按需加 backend flag/压缩。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::ChromCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/pos",
        "7",
        "pos",
        "CompressedField(method=DELTA)",
        "Y",
        "block field",
        "POS 使用 Delta payload：`[original_len][first_value][delta_count][delta_varint...]`。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::PosCompressor::Compress(); src/gvcf/gvcf_encoding.cpp::DeltaEncoder",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/id",
        "8",
        "id",
        "CompressedField(method=MASK or DICTIONARY)",
        "Y",
        "block field",
        "ID 中 `.` 占比达到阈值时用 `Mask('.')`，否则用 Dictionary。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::IdCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/ref",
        "9",
        "ref",
        "CompressedField(method=MASK; specialized 2-bit)",
        "Y",
        "block field",
        "单碱基 A/C/G/T 按 2-bit 打包，非单碱基或非 ACGT 作为 exception。",
        "payload: `[exception_count][delta_indices][exception strings][packed_bases_size][packed_bases]`。",
        "src/gvcf/gvcf_field_compress.cpp::RefCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/alt",
        "10",
        "alt",
        "CompressedField(method=MASK + extra_alts)",
        "Y",
        "block field",
        "第一 ALT 走 Mask，额外 ALT 追加保存。",
        "payload: `Mask(first_alt) + [has_extra:u8][extra_count][record_idx][extra_count][len+bytes...]`。",
        "src/gvcf/gvcf_field_compress.cpp::AltCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/qual",
        "11",
        "qual",
        "CompressedField(method=NONE)",
        "Y",
        "block field",
        "QUAL 直接写 `[count][float32 little-endian bytes...]`，再按需后端压缩。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::QualCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/filter",
        "12",
        "filter",
        "CompressedField(method=RLE/MASK/DICTIONARY)",
        "Y",
        "block field",
        "全相同用 RLE；dominant ratio 达阈值用 Mask；否则 Dictionary。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::FilterCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "known_fields/info_end",
        "13",
        "info_end",
        "CompressedField(method=MASK; inferred exceptions)",
        "if has_end_field",
        "flags bit0",
        "可由下一条 POS 推断的 END 不存值；只存 exception index 和 `END-POS` 差值。",
        "payload: `[exception_count][delta_exception_indices...][end_minus_pos values...]`。",
        "src/gvcf/gvcf_field_compress.cpp::EndCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "sample_fields/gt",
        "14-16",
        "gt_mask / gt_patches / gt_phase",
        "MASK + empty patch wrapper + RLE phase bitmap",
        "Y",
        "block field",
        "GT 字符串以默认 `0/0` 为 dominant 做 Mask；`gt_patches` 仍序列化但当前主要为空 wrapper；相位 bool bitmap 用 RLEByte。",
        "phase payload: `[count][RLEByte(phase_bitmap)]`。",
        "src/gvcf/gvcf_field_compress.cpp::GTCompressor::Compress(); src/gvcf/gvcf_block.cpp::Serialize()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "sample_fields/dp",
        "17",
        "dp",
        "CompressedField(method=RLE or DICTIONARY)",
        "wrapper always; data if present",
        "block field",
        "FORMAT/DP 尝试 RLEInt 与 DictInt，选择序列化后更小的方案。",
        "若 MIN_DP 存在，DP 与 DP-MIN_DP diff 一起由 MinDPCompressor 写出。",
        "src/gvcf/gvcf_field_compress.cpp::DPCompressor::Compress(); MinDPCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "sample_fields/gq",
        "18",
        "gq",
        "CompressedField(method=MASK; PL prediction exceptions)",
        "wrapper always; data if present",
        "block field",
        "FORMAT/GQ 由 PL 预测；预测不匹配的位置保存 exception。",
        "payload: `[count][bitmask_size][RLEByte(predicted_mask)][exception_count][idx_size?][Delta(indices)][exception values...]`。",
        "src/gvcf/gvcf_field_compress.cpp::GQCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "sample_fields/min_dp",
        "19-20",
        "min_dp / dp_min_dp_diff",
        "CompressedField(method=MASK or RLE)",
        "if has_min_dp",
        "flags bit1",
        "MIN_DP 存在时写 `dp` 和 `dp_min_dp_diff`；差值 0 占比较高时用 Mask，否则 RLEInt。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::MinDPCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "sample_fields/pl",
        "21",
        "pl",
        "CompressedField(method=DICTIONARY; optimized v2)",
        "wrapper always; data if present",
        "block field",
        "PL optimized v2：标准三元组 bitmask + PL1/PL2 + exception dictionary + exception indices。",
        "payload starts with `version=2`。",
        "src/gvcf/gvcf_field_compress.cpp::PLCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "sample_fields/ad",
        "22",
        "ad",
        "CompressedField(method=DICTIONARY; PLCompressor shape)",
        "wrapper always; data if present",
        "block field",
        "AD 当前复用 PLCompressor 的矢量整数编码形状。",
        "",
        "src/gvcf/gvcf_field_compress.cpp::ADCompressor::Compress()",
    ),
    (
        "4_FIELD_PAYLOADS_AND_METHODS",
        "unknown_info/unknown_format",
        "23-26",
        "unknown field",
        "name_len + name + CompressedField(method=DICTIONARY)",
        "if present",
        "unordered_map iteration order",
        "未知 INFO/FORMAT 字段按字段名加 CompressedField 写入，字段 payload 使用 GenericFieldCompressor 的 Dictionary。",
        "由于容器是 unordered_map，同类未知字段的写出顺序不保证稳定排序。",
        "src/gvcf/gvcf_block.cpp::Serialize(); src/gvcf/gvcf_field_compress.cpp::GenericFieldCompressor::Compress()",
    ),
]

VCF_HIERARCHY_ROWS = [
    srow("/\nMULTI_SAMPLE_GSC\n多样本 VCF/BCF `.gsc` 文件", "Y", "当前源码实际写盘结构\n无全局 magic/version"),
    srow("TOP_HEADER(1)\n顶层偏移头", "Y", "↘\n文件开头 17 bytes"),
    srow(None, None, None, "mode_type(1)\n文件模式", "Y", "[bool]\ntrue=lossless false=lossy"),
    srow(None, None, None, "other_fields_offset(2)\nother fields 起始偏移", "Y", "[uint64_t]\nmain archive 结束位置"),
    srow(None, None, None, "sdsl_offset(3)\nSDSL 尾区起始偏移", "Y", "[uint64_t]\n4 个 rrr_vector 起点"),
    srow("MAIN_ARCHIVE(2)\n主 archive 区", "Y", "↘\n范围: header 后到 other_fields_offset"),
    srow(None, None, None, "CHUNK_STREAMS(1)\nchunk 边界/偏移表", "Y", "↘"),
    srow(None, None, None, None, None, None, "chunks_streams_size(1)\n边界项数", "Y", "[uint32_t]"),
    srow(None, None, None, None, None, None, "chunk_stream_item(2)\n单个边界项", "Y * N", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "cur_chunk_actual_pos(1)\n累计真实变体数", "Y", "[uint32_t]"),
    srow(None, None, None, None, None, None, None, None, None, "offset(2)\nchunk payload 文件偏移", "Y", "[size_t]\nfinal 文件中 chunk 不带长度前缀"),
    srow(None, None, None, "BASIC_PARAMS(2)\n全局基础参数", "Y", "↘"),
    srow(None, None, None, None, None, None, "ploidy(1)\n样本倍性", "Y", "[uint8_t]"),
    srow(None, None, None, None, None, None, "max_block_rows(2)\nGT row_block 最大变体数", "Y", "[uint32_t]"),
    srow(None, None, None, None, None, None, "max_block_cols(3)\nGT column_block 最大单倍体数", "Y", "[uint32_t]"),
    srow(None, None, None, None, None, None, "vec_len / no_vec / copy_no(4)\nGT 向量统计", "Y", "[uint64_t] * 3"),
    srow(None, None, None, None, None, None, "copy_origin_payload(5)\ncopy origin 差值位流", "Y", "[used_bits_cp:char][size:int32][bytes]"),
    srow(None, None, None, None, None, None, "n_samples(6)\n样本数", "Y", "[uint32_t]"),
    srow(None, None, None, "RANGE_AND_CHROM_INDEX(3)\n范围/染色体索引", "Y", "↘"),
    srow(None, None, None, None, None, None, "chunks_min_pos[](1)\n每个 chunk 最小 POS", "Y", "[uint32 size] + [int64_t] * N"),
    srow(None, None, None, None, None, None, "where_chrom[](2)\n染色体边界表", "Y", "[uint32 count] + items"),
    srow(None, None, None, None, None, None, None, None, None, "where_chrom_item\nchrom -> chunk boundary", "Y * N", "[chrom_size:size_t][chrom bytes][boundary:int/uint32]"),
    srow(None, None, None, "COLUMN_TILING(4)\nGT 列分块元数据", "Y(new) / compat", "↘"),
    srow(None, None, None, None, None, None, "n_col_blocks(1)\n列块数量", "Y(new)", "[uint32_t]"),
    srow(None, None, None, None, None, None, "col_block_item(2)\n列块范围项", "Y(new) * N", "[start_haplotype:uint32][block_size:uint32]"),
    srow(None, None, None, "PERMUTATIONS(5)\n单倍体排列表", "Y", "↘"),
    srow(None, None, None, None, None, None, "permutation_count(1)\n排列条目数", "Y", "[uint32_t]"),
    srow(None, None, None, None, None, None, "permutation_entry_new(2)\n新格式排列项", "conditional", "[row_block_id][col_block_id][data_size][vint bytes]"),
    srow(None, None, None, None, None, None, "permutation_entry_legacy(3)\n旧格式排列项", "compat", "[row_block_id][data_size][vint bytes]"),
    srow(None, None, None, "META(6)\nVCF header/sample 元数据", "Y", "↘"),
    srow(None, None, None, None, None, None, "meta_magic / meta_backend(1)\nmeta codec 标记", "Y(new) / compat", "[uint32_t GSC_META_MAGIC][uint8_t backend]"),
    srow(None, None, None, None, None, None, "comp_v_header(2)\n压缩 header text", "Y", "[uint32 size][backend-compressed bytes]"),
    srow(None, None, None, None, None, None, "comp_v_samples(3)\n压缩 sample names", "Y", "[uint32 size][backend-compressed bytes]"),
    srow(None, None, None, "FIXED_FIELDS_CHUNKS(7)\nfixed-fields chunk payload 串", "Y", "↘\n所有 chunk 顺序拼接"),
    srow(None, None, None, None, None, None, "GSCF_CHUNK(1)\n当前 fixed-fields chunk", "Y", "见下方 GSCF_FIXED_FIELDS_CHUNK\n由 chunks_streams.offset 定位"),
    srow("GSCF_FIXED_FIELDS_CHUNK(3)\nFixed Fields Chunk", "Y", "↘\nmagic=GSCF, version=2"),
    srow(None, None, None, "CHUNK_HEADER(1)\nchunk 头部", "Y", "↘"),
    srow(None, None, None, None, None, None, "magic(1)\nGSCF magic", "Y", "[uint32_t]\n0x46435347"),
    srow(None, None, None, None, None, None, "version(2)\n目录版本", "Y", "[uint32_t]\n当前 writer=2"),
    srow(None, None, None, None, None, None, "total_variants / row_block_count / flags(3-5)\nchunk 统计", "Y", "[uint32_t] * 3"),
    srow(None, None, None, "ROW_BLOCK_DIRECTORY(2)\nrow block 目录", "Y", "↘"),
    srow(None, None, None, None, None, None, "row_block_entry(1)\n单个目录项", "Y * row_block_count", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "variant_count / first_pos / last_pos(1-3)\n范围信息", "Y", "[uint32_t][int64_t][int64_t]"),
    srow(None, None, None, None, None, None, None, None, None, "field off/size pairs(4-10)\nCHROM/POS/ID/REF/ALT/QUAL/GT", "Y", "每字段 [off:uint32][size:uint32]"),
    srow(None, None, None, "DATA_REGION(3)\n字段数据区", "Y", "↘"),
    srow(None, None, None, None, None, None, "fixed field payloads(1)\n固定列字段", "Y", "CHROM/POS/ID/REF/ALT/QUAL\n均为 backend-compressed payload"),
    srow(None, None, None, None, None, None, "GT index payload(2)\nGT 稀疏索引段", "Y", "backend-compressed bytes + 1-byte marker\n不是原始 GT 文本"),
    srow(None, None, None, "LEGACY_COMPAT(4)\n旧 fixed-fields 格式", "compat", "↘"),
    srow(None, None, None, None, None, None, "legacy_no_variants(1)\n旧 chunk 起始字段", "compat", "[uint32_t]"),
    srow(None, None, None, None, None, None, "legacy_field_entry(2)\n旧字段项", "compat", "[uint32 size] + bytes\n顺序 chrom/pos/id/ref/alt/qual/gt"),
    srow("LOSSLESS_PART2(4)\nFILTER/INFO/FORMAT 容器", "lossless only", "↘\n范围 [other_fields_offset, sdsl_offset)"),
    srow(None, None, None, "PAYLOAD_AREA(1)\nstream part 本体区", "Y", "[bytes]\n先写所有 part payload"),
    srow(None, None, None, "NAMED_STREAMS(2)\n命名逻辑流", "Y", "↘"),
    srow(None, None, None, None, None, None, "part2_params(1)\nlossless 参数流", "Y", "[backend_id:uint8] + compressed v_desc"),
    srow(None, None, None, None, None, None, "key_<i>_size / key_<i>_data(2)\n字段 size/data 子流", "Y * key", "FILTER/INFO/FORMAT 每个 key 通常 2 个流"),
    srow(None, None, None, "PART2_PARAMS_PAYLOAD(3)\nv_desc 解压后结构", "Y", "↘\nappend()/append_str() token 流"),
    srow(None, None, None, None, None, None, "actual_variants[](1)\n真实变体数数组", "Y", "[size] + [values]\nappend(uint32_t) tokens"),
    srow(None, None, None, None, None, None, "keys[](2)\nFILTER/INFO/FORMAT key 表", "Y", "key_id, actual_field_id, key_type, BCF type, name"),
    srow(None, None, None, None, None, None, "backend + FMT dictionaries(3)\nFORMAT 特殊 codec 字典", "Y / compat", "backend, fmt_dict_version, fmt_dict_size, blob"),
    srow(None, None, None, "FILE_HANDLE_2_FOOTER(4)\nstream 目录尾部", "Y", "↘"),
    srow(None, None, None, None, None, None, "n_streams / stream_name / part_count(1)\nstream 目录", "Y", "size_t 变长写法 + NUL 字符串"),
    srow(None, None, None, None, None, None, "part.offset / part.size(2)\npart 定位", "Y * part", "相对 part2 payload 起点"),
    srow(None, None, None, None, None, None, "footer_size(3)\nfooter 字节数", "Y", "[fixed 8 bytes at part2 EOF]"),
    srow("GT_ENCODING(5)\n多样本 GT 语义与索引", "Y", "↘"),
    srow(None, None, None, "GT_BITPLANES(1)\n2 bitplane / haplotype allele", "Y", "↘"),
    srow(None, None, None, None, None, None, "MSB/LSB mapping(1)\nGT 4 态", "Y", "00->0, 01->1, 10->., 11->2"),
    srow(None, None, None, "SPARSE_VECTOR_ROW(2)\nunique vector 稀疏行", "Y", "vint_code bytes\n1-position delta list + 0 terminator"),
    srow(None, None, None, "BYTE_LOCAL_XOR(3)\nbyte 内相邻 bit XOR", "Y", "解压由 initialXORLut() 逆变换"),
    srow(None, None, None, "GT_CODEC_MARKER(4)\nGT 段后端标记", "Y", "0=bsc, 1=zstd, 2=brotli\n位于压缩 GT payload 最后 1 byte"),
    srow("SDSL_TAIL(6)\nzero/copy 位图尾区", "Y", "↘\n从 sdsl_offset 开始"),
    srow(None, None, None, "zeros_only[0](1)\n偶数 parity 全零位图", "Y", "sdsl::rrr_vector<>"),
    srow(None, None, None, "zeros_only[1](2)\n奇数 parity 全零位图", "Y", "sdsl::rrr_vector<>"),
    srow(None, None, None, "copy_bit_vector[0](3)\n偶数 parity copy 位图", "Y", "sdsl::rrr_vector<>"),
    srow(None, None, None, "copy_bit_vector[1](4)\n奇数 parity copy 位图", "Y", "sdsl::rrr_vector<>"),
]

GVCF_HIERARCHY_ROWS = [
    srow("/\nNATIVE_GVCF\n单样本 gVCF 压缩文件", "Y", "当前源码实际写盘结构\nmagic=GVCF, version=4"),
    srow("FILE_HEADER(1)\n文件头", "Y", "↘"),
    srow(None, None, None, "magic(1)\n文件魔数", "Y", "[uint32_t]\n0x47564346"),
    srow(None, None, None, "version(2)\n文件版本", "Y", "[uint32_t]\n当前 writer=4"),
    srow(None, None, None, "backend_id(3)\n压缩后端", "Y(V2+)", "[uint8_t]\nbsc/zstd/brotli enum"),
    srow(None, None, None, "num_samples(4)\n样本数", "Y", "[uint32_t]"),
    srow(None, None, None, "HEADER_TEXT(5)\ngVCF header 文本", "Y", "↘"),
    srow(None, None, None, None, None, None, "header_flag(1)\n是否压缩", "Y(V3+)", "[uint8_t]\n1=compressed, 0=raw"),
    srow(None, None, None, None, None, None, "compressed header branch(2)\n压缩 header", "if flag=1", "[original_size:uint32][compressed_size:uint32][bytes]"),
    srow(None, None, None, None, None, None, "raw header branch(3)\n原始 header", "if flag=0 or V1/V2", "[header_size:uint32][raw bytes]"),
    srow(None, None, None, "SAMPLE_NAMES(6)\n样本名列表", "Y", "[name_size:uint32][name bytes] * num_samples"),
    srow(None, None, None, "placeholder(7)\n头部尾占位", "Y", "[uint64_t]\nreader 跳过"),
    srow("BLOCK_AREA(2)\n块区", "Y", "↘\n重复到 footer 前"),
    srow(None, None, None, "BLOCK_ENTRY(1)\n单个压缩块", "Y * num_blocks", "↘"),
    srow(None, None, None, None, None, None, "block_size(1)\n块 payload 字节数", "Y", "[uint32_t]"),
    srow(None, None, None, None, None, None, "block_payload(2)\nCompressedGVCFBlock", "Y", "[bytes]\n内部再带 block magic/version"),
    srow("COMPRESSED_GVCF_BLOCK(3)\n块 payload 内部", "Y", "↘"),
    srow(None, None, None, "BLOCK_HEADER(1)\n块内部头", "Y", "↘"),
    srow(None, None, None, None, None, None, "block magic(1)\n块 magic", "Y", "4 bytes: G V C F"),
    srow(None, None, None, None, None, None, "block version(2)\n块版本", "Y", "[uint8_t]\n当前固定 1"),
    srow(None, None, None, None, None, None, "variant_count / sample_count(3-4)\n块统计", "Y", "[VarUint] * 2"),
    srow(None, None, None, None, None, None, "flags(5)\n字段存在标记", "Y", "[uint8_t]\nbit0=has_end_field, bit1=has_min_dp"),
    srow(None, None, None, "KNOWN_FIELDS(2)\n固定已知字段", "Y", "↘"),
    srow(None, None, None, None, None, None, "POSITION_FIELDS(1)\n位置字段", "Y", "chrom / pos / id\n均为 CompressedField"),
    srow(None, None, None, None, None, None, "SEQUENCE_FIELDS(2)\n序列字段", "Y", "ref / alt\n均为 CompressedField"),
    srow(None, None, None, None, None, None, "QUALITY_FIELDS(3)\n质量字段", "Y", "qual / filter\n均为 CompressedField"),
    srow(None, None, None, None, None, None, "INFO_END(4)\nINFO/END", "flags & 0x01", "CompressedField"),
    srow(None, None, None, None, None, None, "SAMPLE_FIELDS(5)\nFORMAT 样本字段", "Y/conditional", "gt_mask, gt_patches, gt_phase, dp, gq, min_dp, dp_min_dp_diff, pl, ad"),
    srow(None, None, None, "UNKNOWN_FIELDS(3)\n未知 INFO/FORMAT 字段", "Y", "↘"),
    srow(None, None, None, None, None, None, "unknown_info_count + items(1)\n未知 INFO", "Y", "[count:VarUint] + [name_len][name][CompressedField] * N"),
    srow(None, None, None, None, None, None, "unknown_format_count + items(2)\n未知 FORMAT", "Y", "[count:VarUint] + [name_len][name][CompressedField] * N"),
    srow(None, None, None, "COMPRESSED_FIELD_WRAPPER(4)\n字段统一包装", "Y", "↘"),
    srow(None, None, None, None, None, None, "method(1)\n字段压缩方法", "Y", "[uint8_t FieldCompressionMethod]"),
    srow(None, None, None, None, None, None, "original_count(2)\n原始元素数", "Y", "[VarUint]"),
    srow(None, None, None, None, None, None, "data_size(3)\n字段 payload 长度", "Y", "[VarUint]"),
    srow(None, None, None, None, None, None, "data(4)\n字段编码 payload", "Y", "[backend_flag:uint8?] + payload bytes"),
    srow("FIELD_PAYLOADS(4)\n字段级 payload 形状", "as used", "↘"),
    srow(None, None, None, "CHROM(1)\n染色体", "Y", "RLE(string) + backend flag"),
    srow(None, None, None, "POS(2)\n位置", "Y", "Delta + backend flag"),
    srow(None, None, None, "ID(3)\nID", "Y", "Mask('.') 或 Dictionary + backend flag"),
    srow(None, None, None, "REF(4)\n参考序列", "Y", "2-bit A/C/G/T + exception + backend flag"),
    srow(None, None, None, "ALT(5)\n替换序列", "Y", "Mask(first ALT) + extra_alts + backend flag"),
    srow(None, None, None, "QUAL(6)\n质量值", "Y", "[count][float32 little-endian...] + backend flag"),
    srow(None, None, None, "FILTER(7)\n过滤字段", "Y", "RLE / Mask / Dictionary + backend flag"),
    srow(None, None, None, "INFO/END(8)\nEND 范围字段", "if present", "inference exceptions: exception_count + delta indices + END-POS"),
    srow(None, None, None, "GT(9)\nFORMAT/GT", "Y", "gt_mask=Mask(default 0/0), gt_phase=RLEByte bitmap, gt_patches 保留 wrapper"),
    srow(None, None, None, "DP/MIN_DP/GQ(10)\n深度与质量", "if present", "DP=RLE/Dictionary; MIN_DP=DP-MIN_DP diff; GQ=PL prediction exceptions"),
    srow(None, None, None, "PL/AD(11)\n向量整数 FORMAT", "if present", "PL optimized v2; AD 当前复用 PLCompressor 形状"),
    srow("BASIC_ENCODINGS(5)\ngVCF 基础编码器", "as used", "↘"),
    srow(None, None, None, "VarUint / VarInt(1)\n变长整数", "as used", "7-bit continuation; signed uses ZigZag"),
    srow(None, None, None, "RLE(string/int/byte)(2)\n游程编码", "as used", "[original_len][run_count]{value,count}"),
    srow(None, None, None, "Delta(3)\n差分编码", "as used", "[original_len][first_value][delta_count][delta...]"),
    srow(None, None, None, "Mask(string/int)(4)\ndominant + patch", "as used", "dominant + RLEByte bitmask + patch indices + patches"),
    srow(None, None, None, "Dictionary(string/int)(5)\n字典编码", "as used", "dictionary + raw/RLE indices"),
    srow("FILE_FOOTER(6)\n文件尾与范围索引", "Y", "↘\nfooter_offset 指向这里"),
    srow(None, None, None, "num_blocks(1)\n块数量", "Y", "[uint32_t]"),
    srow(None, None, None, "block_offsets[](2)\n块偏移表", "Y", "[uint64_t] * num_blocks\n指向 block_size"),
    srow(None, None, None, "total_variants(3)\n总变体数", "Y", "[uint64_t]"),
    srow(None, None, None, "BLOCK_INDEX(4)\n范围查询索引", "Y(V4)", "↘"),
    srow(None, None, None, None, None, None, "num_indices(1)\n索引项数", "Y(V4)", "[uint32_t]"),
    srow(None, None, None, None, None, None, "block_index_item(2)\n单个索引项", "Y(V4) * N", "[chrom_len][chrom][start_pos][end_pos][file_offset][variant_count]"),
    srow(None, None, None, "footer_offset(5)\nfooter 起始偏移", "Y", "[uint64_t at EOF]\n文件最后 8 字节"),
]


def rows_from_archive(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for path, order, field, encoding, required, desc, note, source in rows:
        out.append((section, path, order, field, encoding, required, "", desc, note, source))
    return out


def rows_from_part2(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for path, field, encoding, required, desc, note, source in rows:
        out.append((section, path, "", field, encoding, required, "", desc, note, source))
    return out


def rows_from_gt(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for category, field, desc, note, source in rows:
        out.append((section, category, "", field, "", "", "", desc, note, source))
    return out


def rows_from_fmt(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for field, codec, shape, note, source in rows:
        out.append((section, "part2 FORMAT special", "", field, codec, "lossless if present", "", shape, note, source))
    return out


def rows_from_gvcf_file(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for row in rows:
        if len(row) == 9:
            path, order, field, encoding, required, version, desc, note, source = row
        else:
            path, order, field, encoding, required, version, desc, source = row
            note = ""
        out.append((section, path, order, field, encoding, required, version, desc, note, source))
    return out


def rows_from_gvcf_block(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for row in rows:
        if len(row) == 8:
            path, order, field, encoding, condition, desc, note, source = row
        else:
            path, order, field, encoding, condition, desc, source = row
            note = ""
        out.append((section, path, order, field, encoding, condition, "", desc, note, source))
    return out


def rows_from_encoding(section: str, rows: list[tuple[str, ...]]) -> list[tuple[str, ...]]:
    out = []
    for name, shape, desc, source in rows:
        out.append((section, "encoding", "", name, shape, "as used", "", desc, "", source))
    return out


def build_vcf_rows() -> list[tuple[str, ...]]:
    rows: list[tuple[str, ...]] = []
    rows.extend(rows_from_archive("1_TOP_HEADER_AND_MAIN_ARCHIVE", BASE.MULTI_ARCHIVE_ROWS))
    rows.extend(rows_from_archive("2_GSCF_FIXED_FIELDS_CHUNK", BASE.FIXED_FIELDS_ROWS))
    rows.extend(rows_from_part2("3_LOSSLESS_PART2", BASE.PART2_ROWS))
    rows.extend(rows_from_gt("4_GT_ENCODING_AND_SDSL", BASE.GT_ROWS))
    rows.extend(rows_from_fmt("5_FORMAT_SPECIAL_CODECS", BASE.FMT_SPECIAL_ROWS))
    rows.extend(VCF_AUX_ROWS)
    return rows


def build_gvcf_rows() -> list[tuple[str, ...]]:
    rows: list[tuple[str, ...]] = []
    rows.extend(rows_from_gvcf_file("1_FILE_HEADER_BLOCK_AREA_FOOTER", BASE.GVCF_FILE_ROWS))
    rows.extend(rows_from_gvcf_block("2_COMPRESSED_GVCF_BLOCK", BASE.GVCF_BLOCK_ROWS))
    rows.extend(GVCF_FIELD_ROWS)
    rows.extend(rows_from_encoding("5_BASIC_ENCODINGS", BASE.GVCF_ENCODING_ROWS))
    return rows


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
STRUCTURE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CFE2F3")
STRUCTURE_SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="EAF3FB")
SECTION_FILLS = {
    "1": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "2": PatternFill(fill_type="solid", fgColor="EAF3FB"),
    "3": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "4": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "5": PatternFill(fill_type="solid", fgColor="EADCF8"),
    "6": PatternFill(fill_type="solid", fgColor="E7E6E6"),
}
HEADER_FONT = Font(bold=True)
WRAP = Alignment(vertical="top", horizontal="left", wrap_text=True)


def autosize(ws, columns: list[str], rows: list[tuple[str, ...]]) -> None:
    widths = {
        1: 28,
        2: 30,
        3: 10,
        4: 28,
        5: 34,
        6: 16,
        7: 18,
        8: 58,
        9: 54,
        10: 58,
    }
    for idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(idx, 20)


def write_table_sheet(wb: Workbook, title: str, columns: list[str], rows: list[tuple[str, ...]]) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    for row_idx, values in enumerate(rows, start=2):
        section = str(values[0])
        fill = SECTION_FILLS.get(section[:1])
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            if fill:
                cell.fill = fill

    autosize(ws, columns, rows)
    ws.row_dimensions[1].height = 32


def write_structure_sheet(wb: Workbook, title: str, rows: list[tuple[str | None, ...]]) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A3"

    for rng in ("A1:C1", "D1:F1", "G1:I1", "J1:L1"):
        ws.merge_cells(rng)

    for col_idx, value in enumerate(STRUCTURE_HEADER_ROW, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = STRUCTURE_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    for col_idx, value in enumerate(STRUCTURE_LABEL_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = STRUCTURE_SUBHEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    level_fills = {
        0: PatternFill(fill_type="solid", fgColor="D9EAD3"),
        3: PatternFill(fill_type="solid", fgColor="FFF2CC"),
        6: PatternFill(fill_type="solid", fgColor="DDEBF7"),
        9: PatternFill(fill_type="solid", fgColor="FCE4D6"),
    }

    for row_idx, row in enumerate(rows, start=3):
        non_empty_groups = [idx for idx in (0, 3, 6, 9) if row[idx] is not None]
        fill = level_fills.get(non_empty_groups[0]) if non_empty_groups else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            if fill:
                cell.fill = fill
                if col_idx in (1, 4, 7, 10) and value is not None:
                    cell.font = HEADER_FONT

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42
    for col_idx, width in enumerate(STRUCTURE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_info_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "说明"
    columns = ["对象", "说明"]
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for row_idx, row in enumerate(INFO_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"


def write_constants_sheet(wb: Workbook) -> None:
    write_table_sheet(wb, "常量枚举", BASE.CONSTANT_COLUMNS, BASE.CONSTANT_ROWS)


def main() -> None:
    wb = Workbook()
    write_info_sheet(wb)
    write_structure_sheet(wb, "VCF层次结构", VCF_HIERARCHY_ROWS)
    write_structure_sheet(wb, "gVCF层次结构", GVCF_HIERARCHY_ROWS)
    write_table_sheet(wb, "VCF_gsc二进制结构", COMMON_COLUMNS, build_vcf_rows())
    write_table_sheet(wb, "gVCF_native二进制结构", COMMON_COLUMNS, build_gvcf_rows())
    write_constants_sheet(wb)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
