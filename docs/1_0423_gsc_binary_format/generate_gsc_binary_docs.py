from __future__ import annotations

from pathlib import Path
from textwrap import dedent

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUT_MD = ROOT / "GSC_binary_format_260423.md"
OUT_XLSX = ROOT / "GSC_binary_format_260423.xlsx"
OUT_OVERVIEW_XLSX = ROOT / "GSC_VCF_gVCF二进制总体结构表-260423.xlsx"
OUT_GSC_LOGIC_SVG = ROOT / "00_GSC_logic_structure.svg"


INFO_COLUMNS = ["对象", "说明"]
INFO_ROWS = [
    ("文档范围", "覆盖当前仓库里的两条实际 on-disk 二进制格式：多样本 VCF/BCF -> `.gsc`，以及单样本 `gsc gvcf` 输出的 `GVCF` native 文件。"),
    ("不覆盖范围", "不把 `docs/0_avsStructure/基因比对压缩封装格式-241227.xlsx` 里的 `avsg` 外层封装当成当前事实格式；这里只整理现有源码真实写出的字节布局。"),
    ("参考目录", "参考了 `docs/0_avsStructure/` 里的现有说明、结构表和汇报稿，但所有字段顺序、宽度和条件均以源码 writer/reader 为准。"),
    ("多样本主线", "核心入口是 `src/compressor.cpp::writeCompressFlie()` 与 `src/decompression_reader.cpp::OpenReading()/readFixedFields()`。"),
    ("gVCF 主线", "核心入口是 `src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/CompressAndWriteBlock()/WriteFileFooter()` 与 `ReadFileHeader()/ReadBlock()`。"),
    ("0423 输出目录", "本次整理结果放在 `docs/0423_gsc_binary_format/`，包含 1 份 Markdown 总文档、1 份详细字段工作簿和 1 份总体结构工作簿。"),
    ("当前 writer 状态", "多样本 `.gsc` 当前 writer 使用 `GSCF V2` fixed-fields chunk；gVCF 当前 writer 输出 `GVCF_FILE_VERSION = 4`。"),
    ("字节序与 ABI", "多样本 `.gsc` 主 archive 直接写入 `bool` 与 `size_t`，默认依赖 64-bit little-endian ABI；part2 footer 里的 `footer_size` 也是固定 8 字节落盘。"),
    ("字符串/数字辅助编码", "多样本 fixed fields、meta 和 part2_params 大量复用 `append_str()/read_str()` 与 `append()/read()`：字符串是 `\\0` 结尾，数字是十进制 ASCII 文本再补 `\\0`，不是定长整数。"),
    ("交付目标", "让 `.gsc` / `GVCF` 的文件级布局、区段、字段、payload 包装、兼容分支和关键常量，都能在 Markdown 与 Excel 里一一对照。"),
]


OVERVIEW_COLUMNS = ["格式", "一级区段", "二级区段", "字段/条目", "类型/编码", "是否必选", "说明", "来源"]
OVERVIEW_ROWS = [
    ("多样本 `.gsc`", "TOP_HEADER", "-", "`mode_type`, `other_fields_offset`, `sdsl_offset`", "bool + uint64 + uint64", "Y", "文件开头无全局 magic/version，先写 17 字节顶层头。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("多样本 `.gsc`", "MAIN_ARCHIVE", "chunk 边界 / GT 索引 / meta / fixed-fields chunks", "`chunks_streams[]`, GT copy/zero 索引、meta、chunk payloads", "复合顺序布局", "Y", "主 archive 从顶层头后开始，一直到 `other_fields_offset`。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("多样本 `.gsc`", "LOSSLESS_PART2", "other fields 多流容器", "`key_i_size`, `key_i_data`, `part2_params`, footer", "File_Handle_2 容器", "lossless only", "仅 `mode_type=true` 时存在；区间是 `[other_fields_offset, sdsl_offset)`。", "src/file_handle.cpp; src/compressor.cpp; src/decompression_reader.cpp::OpenReadingPart2()"),
    ("多样本 `.gsc`", "SDSL_TAIL", "-", "`zeros[2]`, `copies[2]`", "4 x `sdsl::rrr_vector<>`", "Y", "解压时先跳到 `sdsl_offset` 加载。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("单样本 `GVCF`", "FILE_HEADER", "-", "`magic`, `version`, backend, header text, sample names", "定长整数 + 条件分支", "Y", "顶层带 `GVCF` magic 和文件版本。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader(); ReadFileHeader()"),
    ("单样本 `GVCF`", "BLOCK_AREA", "逐块 envelope", "`block_size` + `CompressedGVCFBlock`", "`uint32` + 变长 payload", "Y", "每个 block 先写长度，再写块内序列化字节流。", "src/gvcf/gvcf_compressor.cpp::CompressAndWriteBlock(); ReadBlock()"),
    ("单样本 `GVCF`", "FILE_FOOTER", "块偏移与范围索引", "`num_blocks`, `block_offsets[]`, `total_variants`, `block_indices[]`, `footer_offset`", "定长整数 + 重复项", "Y", "V4 footer 含范围查询索引。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter(); ReadFileHeader(); GVCFQueryer::Open()"),
]


MULTI_ARCHIVE_COLUMNS = ["层级路径", "顺序", "字段", "类型/编码", "当前 writer", "说明", "备注", "来源"]
MULTI_ARCHIVE_ROWS = [
    ("top_header", "1", "mode_type", "bool", "Y", "`true=lossless`，`false=lossy`。", "CLI 解压模式必须和文件模式匹配。", "src/compressor.cpp::writeCompressFlie(); src/decompressor.cpp::decompressProcess()"),
    ("top_header", "2", "other_fields_offset", "uint64_t", "Y", "lossless part2 区域起始偏移。", "lossy 文件通常与 `sdsl_offset` 相同。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("top_header", "3", "sdsl_offset", "uint64_t", "Y", "SDSL 尾区起始偏移。", "从这里开始连续读取 4 个 `rrr_vector`。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/chunk_streams", "1", "chunks_streams_size", "uint32_t", "Y", "chunk 边界表项数。", "每项对应一个边界，而不是 chunk 内长度。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/chunk_streams", "2", "chunk_stream_item", "`uint32_t cur_chunk_actual_pos` + `size_t offset`", "Y", "记录累计真实变体数和 chunk payload 的绝对文件偏移。", "final `.gsc` 中每个 chunk 前没有长度前缀。", "src/utils.h; src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "3", "ploidy", "uint8_t", "Y", "样本倍性。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "4", "max_block_rows", "uint32_t", "Y", "GT row block 最大变体数。", "0 会在 reader 侧回退为总单倍体数。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "5", "max_block_cols", "uint32_t", "Y", "GT column block 最大单倍体数。", "0 或 >= 总单倍体时等价于不切列。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "6", "vec_len", "uint64_t", "Y", "legacy 单列块模式下单个 GT 向量的字节长度。", "切列时 reader 会优先使用 `n_col_blocks` 元数据。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "7", "no_vec", "uint64_t", "Y", "GT 向量总数。", "通常等于 `2 * 变体数`。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "8", "copy_no", "uint64_t", "Y", "被标记为 copy 的向量数。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "9", "used_bits_cp", "char", "Y", "copy-origin 差值打包所需 bit 数。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "10", "bm_comp_cp_size", "int32_t", "Y", "copy-origin bitstream 大小。", "reader 侧按 `int` 读取。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "11", "bm_comp_copy_orgl_id", "bytes", "Y", "copy-origin 差值的位级打包数据。", "由 `CBitMemory` 写出。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "12", "n_samples", "uint32_t", "Y", "样本数。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "13", "chunks_min_pos_size", "uint32_t", "Y", "每个 chunk 的最小 POS 数组长度。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "14", "chunks_min_pos[]", "int64_t[]", "Y", "每个 chunk 的最小 POS。", "用于范围筛选。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "15", "where_chrom_size", "uint32_t", "Y", "染色体边界项个数。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/basic", "16", "where_chrom_item", "`size_t chrom_size` + `chrom bytes` + `int`", "Y", "记录 `chrom -> chunk boundary`。", "writer 用 `int` 写 `elem.second`，reader 按 `uint32_t` 读。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/column_tiling", "17", "n_col_blocks", "uint32_t", "Y", "GT 列分块数量。", "旧文件可能没有这段；reader 会回退为 1。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/column_tiling", "18", "col_block_range_item", "`start_haplotype:uint32_t` + `block_size:uint32_t`", "Y", "描述每个列块覆盖的单倍体范围。", "reader 额外计算 `vec_len=(size+7)/8`。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/permutation", "19", "permutation_count", "uint32_t", "Y", "排列条目数。", "新旧格式共用同一个 count。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/permutation", "20", "permutation_entry(new)", "`row_block_id:uint32_t` + `col_block_id:uint32_t` + `data_size:uint32_t` + `bytes`", "conditional", "列分块新格式的排列条目。", "`useLegacyPath=false` 时 reader 按这个布局读取。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/permutation", "20", "permutation_entry(legacy)", "`row_block_id:uint32_t` + `data_size:uint32_t` + `bytes`", "conditional", "旧格式只按 row block 写排列。", "reader 会回退到 legacy map。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/meta", "21", "meta_magic", "uint32_t", "Y", "`GSC_META_MAGIC`，用于说明后面跟着 meta backend 标记。", "老文件可能没有这一段。", "src/defs.h; src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/meta", "22", "meta_backend", "uint8_t", "Y", "meta/header/sample 使用的 backend id。", "reader 先据此切换 backend，再解压 meta。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/meta", "23", "comp_v_header_size", "uint32_t", "Y", "压缩后 header payload 大小。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/meta", "24", "comp_v_header", "bytes", "Y", "backend 压缩后的 `all_v_header`。", "`all_v_header` 本体是 1 个 `append_str(header)`。", "src/compressor.cpp::compress_meta(); src/decompression_reader.cpp::decompress_meta()"),
    ("main_archive/meta", "25", "comp_v_samples_size", "uint32_t", "Y", "压缩后 sample 列表 payload 大小。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("main_archive/meta", "26", "comp_v_samples", "bytes", "Y", "backend 压缩后的 `all_v_samples`。", "`all_v_samples` 本体是重复的 `append_str(sample_name)`。", "src/compressor.cpp::compress_meta(); src/decompression_reader.cpp::decompress_meta()"),
    ("main_archive/chunks", "27", "fixed-fields chunk payloads", "bytes (concatenated)", "Y", "所有 chunk payload 顺序拼接。", "最终文件里没有 chunk-size 前缀，offset 靠 `chunks_streams` 回填。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::setStartChunk()/readFixedFields()"),
    ("sdsl_tail", "28", "zeros_only[0]", "`sdsl::rrr_vector<>`", "Y", "偶数 parity 的全零向量位图。", "从 `sdsl_offset` 起连续序列化。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("sdsl_tail", "29", "zeros_only[1]", "`sdsl::rrr_vector<>`", "Y", "奇数 parity 的全零向量位图。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("sdsl_tail", "30", "copy_bit_vector[0]", "`sdsl::rrr_vector<>`", "Y", "偶数 parity 的 copy 向量位图。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    ("sdsl_tail", "31", "copy_bit_vector[1]", "`sdsl::rrr_vector<>`", "Y", "奇数 parity 的 copy 向量位图。", "/", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
]


FIXED_FIELDS_COLUMNS = ["层级路径", "顺序", "字段", "类型/编码", "当前 writer", "说明", "备注", "来源"]
FIXED_FIELDS_ROWS = [
    ("chunk_wrapper(temp only)", "0", "payload_size", "uint64_t", "temp only", "临时文件里的 chunk payload 前缀长度。", "写最终 `.gsc` 时这个前缀不会保留。", "src/compressor.cpp::compressFixedFieldsChunk(); src/compressor.cpp::writeCompressFlie()"),
    ("chunk_header", "1", "magic", "uint32_t", "Y", "`GSC_FIXED_FIELDS_RB_MAGIC = 0x46435347`。", "little-endian 下可视为 `GSCF`。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("chunk_header", "2", "version", "uint32_t", "Y", "当前 writer 输出 `GSC_FIXED_FIELDS_RB_VERSION_V2 = 2`。", "reader 同时兼容 V1 与 legacy 平铺格式。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("chunk_header", "3", "total_variants", "uint32_t", "Y", "该 chunk 覆盖的总变体数。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("chunk_header", "4", "row_block_count", "uint32_t", "Y", "chunk 内 row block 数。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("chunk_header", "5", "flags", "uint32_t", "Y", "当前 writer 置 0。", "预留字段。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("chunk_header(v1 compat)", "6", "fixed_fields_gt_off", "uint32_t", "V1 only", "V1 头部里的 chunk-level GT payload 偏移。", "V2 改为每个 row block 自带 `gt_off/gt_size`。", "src/decompression_reader.cpp::readFixedFields()"),
    ("chunk_header(v1 compat)", "7", "fixed_fields_gt_size", "uint32_t", "V1 only", "V1 头部里的 chunk-level GT payload 大小。", "/", "src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "1", "variant_count", "uint32_t", "Y", "该 row block 的变体数。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "2", "first_pos", "int64_t", "Y", "row block 第一个变体的 POS。", "范围查询按它和 `last_pos` 剪裁。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::DecoderByRange()"),
    ("row_block_dir", "3", "last_pos", "int64_t", "Y", "row block 最后一个变体的 POS。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::DecoderByRange()"),
    ("row_block_dir", "4", "chrom_off / chrom_size", "uint32_t + uint32_t", "Y", "`chrom` payload 在 chunk 内的位置与长度。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "5", "pos_off / pos_size", "uint32_t + uint32_t", "Y", "`pos` payload 在 chunk 内的位置与长度。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "6", "id_off / id_size", "uint32_t + uint32_t", "Y", "`id` payload 在 chunk 内的位置与长度。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "7", "ref_off / ref_size", "uint32_t + uint32_t", "Y", "`ref` payload 在 chunk 内的位置与长度。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "8", "alt_off / alt_size", "uint32_t + uint32_t", "Y", "`alt` payload 在 chunk 内的位置与长度。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "9", "qual_off / qual_size", "uint32_t + uint32_t", "Y", "`qual` payload 在 chunk 内的位置与长度。", "/", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("row_block_dir", "10", "gt_off / gt_size", "uint32_t + uint32_t", "Y (V2)", "该 row block 的 GT index payload 在 chunk 内的位置与长度。", "V2 才有；writer 当前就是 V2。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("data_region", "11", "row_block payload order", "`chrom -> pos -> id -> ref -> alt -> qual -> gt`", "Y", "data region 按 row block 顺序紧挨着写。", "reader 实际依赖目录里的 `off/size`，不依赖硬编码顺序。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    ("payload_semantics", "12", "chrom/id/alt/qual/ref", "压缩后 bytes；解压后是重复的 `append_str()` 字符串", "Y", "每个值都以 `\\0` 结尾。", "解码时靠 `read_str()` 顺序取回。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.cpp; src/decompression_reader.cpp"),
    ("payload_semantics", "13", "pos", "压缩后 bytes；解压后是重复的 `append(int64_t)` 结果", "Y", "每个差分值先转十进制 ASCII，再补 `\\0`。", "reader 用 `read()` 取出后累加到 `prev_pos`。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.h; src/decompression_reader.cpp"),
    ("payload_semantics", "14", "gt", "压缩后 bytes + 尾部 marker", "Y", "GT 本体是 row block 级 GT index payload，而不是原始 GT 文本。", "writer 会在每段压缩 GT payload 最后 `push_back(marker)`。", "src/compressor.cpp::compressFixedFieldsChunkToPayload(); src/decompression_reader.cpp"),
    ("legacy_chunk", "15", "legacy no_variants", "uint32_t", "compat", "旧文件 fixed-fields chunk 起点不是 `GSCF`，而是直接从 `no_variants` 开始。", "reader 仍保留 legacy 解码路径。", "src/compressor.cpp::writeTempFlie(); src/decompression_reader.cpp::readFixedFields()"),
    ("legacy_chunk", "16", "legacy field entry", "`uint32_t size` + `bytes`", "compat", "旧格式依次写 `chrom/pos/id/ref/alt/qual/gt`。", "没有 row block 目录，不能做精细范围裁剪。", "src/compressor.cpp::writeTempFlie(); src/decompression_reader.cpp::readFixedFields()"),
]


PART2_COLUMNS = ["层级路径", "字段/流", "类型/编码", "当前 writer", "说明", "备注", "来源"]
PART2_ROWS = [
    ("part2_region", "region", "`[other_fields_offset, sdsl_offset)`", "lossless only", "多流 other-fields 容器整体区间。", "由 `File_Handle_2` 管理。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReadingPart2()"),
    ("part2_payload", "stream payload area", "bytes", "Y", "所有 stream part 先顺序写 payload，本体没有全局 magic。", "footer 在 payload 后面。", "src/file_handle.cpp::AddPartComplete(); serialize()"),
    ("part2_stream_names", "part2_params", "string stream_name", "Y", "记录 `actual_variants`、keys、backend、FMT dictionaries。", "读侧先用这个流恢复 part2 元信息。", "src/compressor.cpp; src/decompression_reader.cpp::OpenReadingPart2()"),
    ("part2_stream_names", "`key_<i>_size`", "string stream_name", "Y", "第 `i` 个 FILTER/INFO/FORMAT 字段的 size 子流。", "一个字段通常拆成 size/data 两个流。", "src/compression_reader.cpp::InitVarinats(); src/decompression_reader.cpp"),
    ("part2_stream_names", "`key_<i>_data`", "string stream_name", "Y", "第 `i` 个字段的 data 子流。", "/", "src/compression_reader.cpp::InitVarinats(); src/decompression_reader.cpp"),
    ("part2_footer", "n_streams", "`File_Handle_2::Write(size_t)`", "Y", "stream 个数。", "`Write(size_t)` 不是定长整数，而是 `[no_bytes:u8][big-endian bytes]`。", "src/file_handle.cpp::serialize(); Write(size_t); read(size_t)"),
    ("part2_footer", "stream_name", "零结尾字符串", "Y", "每个 stream 的逻辑名字。", "`Write(string)` = 原文 bytes + `\\0`。", "src/file_handle.cpp::serialize(); Write(string)"),
    ("part2_footer", "part_count", "`File_Handle_2::Write(size_t)`", "Y", "该 stream 包含的 part 数。", "/", "src/file_handle.cpp::serialize(); read(size_t)"),
    ("part2_footer", "part.offset", "`File_Handle_2::Write(size_t)`", "Y", "part 在 payload 区里的偏移。", "相对于 part2 payload 起点，而不是相对于整个 `.gsc` 文件。", "src/file_handle.cpp::serialize(); deserialize()"),
    ("part2_footer", "part.size", "`File_Handle_2::Write(size_t)`", "Y", "part 字节长度。", "/", "src/file_handle.cpp::serialize(); deserialize()"),
    ("part2_footer", "footer_size", "fixed 8 bytes", "Y", "footer 自身字节数。", "`WriteFixed()` 直接写 8 字节原始内存，依赖 little-endian / 64-bit ABI。", "src/file_handle.cpp::serialize(); deserialize()"),
    ("part2_params transport", "backend_id prefix", "uint8_t", "Y", "`part2_params` 这个 part 的第 1 字节，表示其压缩 backend。", "reader 先读它，再解压 `v_desc`。", "src/compressor.cpp; src/decompression_reader.cpp::OpenReadingPart2()"),
    ("part2_params transport", "compressed v_desc", "bytes", "Y", "backend 压缩后的 `v_desc`。", "`v_desc` 内部大量字段不是定长整数，而是 `append()/append_str()` 生成的 token 流。", "src/compressor.cpp; src/decompression_reader.cpp::OpenReadingPart2()"),
    ("part2_params payload", "actual_variants_size", "`append(uint32_t)`", "Y", "chunk 级 `actual_variants[]` 的长度。", "实际写法是十进制 ASCII + `\\0`。", "src/compressor.cpp; src/utils.h"),
    ("part2_params payload", "actual_variants[]", "repeated `append(uint32_t)`", "Y", "每个 chunk 对应的真实变体数。", "多等位位点拆分会导致它与原始输入记录数不一定 1:1。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "no_keys", "`append(uint32_t)`", "Y", "FILTER/INFO/FORMAT key 总数。", "/", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "key_gt_id", "`append(int)`", "Y", "GT 在 key 表里的位置。", "/", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "key entry", "`key_id`, `actual_field_id`, `keys_type`, `type`, `name`", "Y", "逐个保存字段元信息。", "前四个值走 `append()`，`name` 走 `append_str()`。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "params.backend", "`append(uint32_t)`", "Y", "part2 内部 special codec 也会记录 backend。", "和 transport 前缀是两层不同语义。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "fmt_dict_version", "`append(uint32_t)`", "Y", "当前固定写 1。", "旧文件可能没有这个尾巴。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "fmt_dict_size", "`append(uint32_t)`", "Y", "FMT dictionaries blob 大小。", "/", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("part2_params payload", "fmt_dict_blob", "bytes", "Y", "AD / PL / PID dictionaries 序列化字节。", "reader 先恢复它，再解码 special FMT codec。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("key_i_size payload", "size stream", "backend-compressed bytes", "Y", "解压后是 `uint32_t` 数组，用于切分每条记录的字段大小。", "reader 直接 `copy_n` 到 `vector<uint32_t>`。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    ("key_i_data payload", "data stream", "backend-compressed bytes", "Y", "非特殊字段按类型压缩；INT 默认先做按 byte 位面转置。", "特殊 FMT 字段会带自己的一层 `codec_id` / record marker。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
]


GT_COLUMNS = ["类别", "字段/取值", "说明", "备注", "来源"]
GT_ROWS = [
    ("GT 两条 bitplane", "MSB/LSB = `00`", "恢复为字符 `0`。", "当前 LUT 固定语义。", "src/compression_reader.cpp::addVariant(); src/decompressor.cpp::initialLut()"),
    ("GT 两条 bitplane", "MSB/LSB = `01`", "恢复为字符 `1`。", "/", "src/compression_reader.cpp::addVariant(); src/decompressor.cpp::initialLut()"),
    ("GT 两条 bitplane", "MSB/LSB = `10`", "恢复为字符 `.`（missing）。", "/", "src/compression_reader.cpp::addVariant(); src/decompressor.cpp::initialLut()"),
    ("GT 两条 bitplane", "MSB/LSB = `11`", "恢复为字符 `2`。", "表示 allele 2。", "src/compression_reader.cpp::addVariant(); src/decompressor.cpp::initialLut()"),
    ("GT row 编码", "`delta list + 0 terminator`", "每条 unique GT 向量存 1 的位置 delta 列表，行尾追加 `0`。", "整数数组再交给 `vint_code::EncodeArray()`。", "src/block_processing.cpp; src/decompressor.cpp::decoded_vector_row()"),
    ("GT 变换域", "byte 内相邻 bit XOR", "稀疏化前的差分变换。", "解压侧用 `initialXORLut()` 逆变换。", "src/decompressor.cpp::initialXORLut(); AGENTS.md 说明"),
    ("GT codec marker", "`0`", "GT payload 使用 BSC。", "marker 放在压缩 GT payload 最后 1 byte。", "src/compressor.cpp::compressFixedFieldsChunkToPayload(); src/decompression_reader.cpp"),
    ("GT codec marker", "`1`", "GT payload 使用 Zstd。", "/", "src/compressor.cpp::compressFixedFieldsChunkToPayload(); src/decompression_reader.cpp"),
    ("GT codec marker", "`2`", "GT payload 使用 Brotli。", "/", "src/compressor.cpp::compressFixedFieldsChunkToPayload(); src/decompression_reader.cpp"),
    ("GT zero/copy 全局索引", "`zeros_only_bit_vector[2]`", "全零向量位图。", "最终落到 SDSL tail。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp"),
    ("GT zero/copy 全局索引", "`copy_bit_vector[2]`", "copy 向量位图。", "copy origin 差值另存在 `bm_comp_copy_orgl_id`。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp"),
    ("切列注意事项", "only last col block carries `variant_desc`", "column tiling 下只有最后一个列块携带 `v_vcf_data_compress`。", "压缩线程不能默认每个列块都带 fixed fields 描述。", "src/compression_reader.cpp::addVariant(); AGENTS.md 说明"),
]


FMT_SPECIAL_COLUMNS = ["字段", "流类型 / codec_id", "on-disk 形状", "备注", "来源"]
FMT_SPECIAL_ROWS = [
    ("DP", "record_codec `0` / `1`", "`0`: 原始 `int32[n_samples]`；`1`: `[raw_count][(delta_pos,val_enc)...][exc_count][(delta_pos,val_enc)...]`，依赖 AD 恢复预测值。", "整个字段 data stream 外面还会再包一层 stream 级 `codec_id=1`。", "src/compression_reader.cpp; src/compressor.cpp; src/decompression_reader.cpp"),
    ("MIN_DP", "record_codec `0` / `1`", "`0`: 原始 `int32[n_samples]`；`1`: `[exc_count][(delta_pos,val_enc)...]`，依赖 DP 恢复默认值。", "同属 record-wise byte stream。", "src/compression_reader.cpp; src/decompression_reader.cpp"),
    ("GQ", "record_codec `0` / `1`", "`0`: 原始 `int32[n_samples]`；`1`: `[raw_count][(delta_pos,val_enc)...][exc_count][(delta_pos,val_enc)...]`，依赖 PL 预测。", "同属 record-wise byte stream。", "src/compression_reader.cpp; src/decompression_reader.cpp"),
    ("PGT", "stream codec_id `0` / `1`", "`0`: legacy raw bytes；`1`: 稀疏格式，每条记录保存非缺失样本位置和对应 stride bytes。", "字符串型 FORMAT 字段。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("PID", "stream codec_id `0` / `1` / `2`", "`0`: legacy raw bytes；`1`: 稀疏原始字符串；`2`: 稀疏 + dictionary tag。", "codec_id=2 依赖 `fmt_dict_blob`。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("AD", "stream codec_id `0` / `5` / `7`", "`0`: legacy transposed INT bytes；`5`: `Δsum` 与 fallback 分流；`7`: 在 `5` 基础上增加 biallelic alt nibble/escape 子流。", "writer 默认优先用 `7`，失败时回退 `0`。", "src/compressor.cpp; src/decompression_reader.cpp"),
    ("PL", "stream codec_id `0` / `6`", "`0`: legacy transposed INT bytes；`6`: 特化的 tip2/split payload 编码。", "reader 对 `codec_id=6` 做额外归一化还原。", "src/compressor.cpp; src/decompression_reader.cpp"),
]


GVCF_FILE_COLUMNS = ["层级路径", "顺序", "字段", "类型/编码", "当前 writer", "条件/版本", "说明", "来源"]
GVCF_FILE_ROWS = [
    ("file_header", "1", "magic", "uint32_t", "Y", "all versions", "`GVCF_FILE_MAGIC = 0x47564346`。", "src/gvcf/gvcf_compressor.h; src/gvcf/gvcf_compressor.cpp"),
    ("file_header", "2", "version", "uint32_t", "Y", "writer 当前写 4", "文件版本。", "reader 当前支持到 V4。", "src/gvcf/gvcf_compressor.h; WriteFileHeader(); ReadFileHeader()"),
    ("file_header", "3", "backend_id", "uint8_t", "Y", "V2+", "块内 header / block payload 的压缩 backend。", "V1 默认按 zstd 兼容。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader(); ReadFileHeader()"),
    ("file_header", "4", "num_samples", "uint32_t", "Y", "all versions", "样本数。", "当前 gVCF 主线通常是单样本，但格式字段仍保留计数。", "src/gvcf/gvcf_compressor.cpp"),
    ("file_header", "5", "header_flag", "uint8_t", "Y", "V3+", "`1` 表示 header 文本被压缩，`0` 表示原样写。", "/", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader(); ReadFileHeader()"),
    ("file_header", "6", "header payload", "conditional", "Y", "V3+", "若 `header_flag=1`：`original_size:uint32 + compressed_size:uint32 + data`；若 `0`：`header_size:uint32 + raw bytes`。", "V1/V2 只有 uncompressed header 分支。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader(); ReadFileHeader()"),
    ("file_header", "7", "sample_names[]", "`repeat [name_size:uint32][name bytes]`", "Y", "all versions", "样本名列表。", "数量由 `num_samples` 决定。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader(); ReadFileHeader()"),
    ("file_header", "8", "placeholder", "uint64_t", "Y", "all versions", "头部尾端保留的 8 字节占位。", "reader 直接跳过，不参与逻辑。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader(); ReadFileHeader()"),
    ("block_area", "9", "block_size", "uint32_t", "Y", "all versions", "单个块序列化字节数。", "后面紧跟 `CompressedGVCFBlock`。", "src/gvcf/gvcf_compressor.cpp::CompressAndWriteBlock(); ReadBlock()"),
    ("block_area", "10", "block_payload", "bytes", "Y", "all versions", "`CompressedGVCFBlock::Serialize()` 输出。", "块内自己再带一个 4-byte `GVCF` magic 和 1-byte block version。", "src/gvcf/gvcf_block.cpp"),
    ("file_footer", "11", "num_blocks", "uint32_t", "Y", "all versions", "block 数。", "footer 开始处第 1 个字段。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter(); ReadFileHeader()"),
    ("file_footer", "12", "block_offsets[]", "uint64_t[]", "Y", "all versions", "每个 block 的文件偏移。", "offset 指向对应 `block_size` 字段。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter(); ReadFileHeader()"),
    ("file_footer", "13", "total_variants", "uint64_t", "Y", "all versions", "总变体数。", "/", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter(); ReadFileHeader()"),
    ("file_footer", "14", "num_indices", "uint32_t", "Y", "V4", "范围查询索引条目数。", "V4 才有。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter(); ReadFileHeader(); GVCFQueryer::Open()"),
    ("file_footer", "15", "block_index_item", "`chrom_len:uint32 + chrom + start_pos:uint64 + end_pos:uint64 + file_offset:uint64 + variant_count:uint32`", "Y", "V4", "每个块的染色体和区间。", "Queryer 先靠它筛候选块。", "src/gvcf/gvcf_compressor.h::BlockIndex; WriteFileFooter(); GVCFQueryer::Open()"),
    ("file_footer", "16", "footer_offset", "uint64_t", "Y", "all versions", "文件尾最后 8 字节，指向 footer 起点。", "reader 先 seek 到文件尾回读它。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter(); ReadFileHeader(); GVCFQueryer::Open()"),
    ("version_compat", "-", "V1", "compat note", "reader only", "V1", "无 backend id、无压缩 header 标志、无 block index。", "/", "src/gvcf/gvcf_compressor.cpp::ReadFileHeader()"),
    ("version_compat", "-", "V2", "compat note", "reader only", "V2", "新增 backend id。", "/", "src/gvcf/gvcf_compressor.cpp::ReadFileHeader()"),
    ("version_compat", "-", "V3", "compat note", "reader only", "V3", "header 支持 `header_flag` + 可选压缩。", "/", "src/gvcf/gvcf_compressor.cpp::ReadFileHeader()"),
    ("version_compat", "-", "V4", "current writer", "Y", "V4", "新增 `block_indices[]`，支持范围查询。", "当前 writer 固定写 V4。", "src/gvcf/gvcf_compressor.h; src/gvcf/gvcf_compressor.cpp"),
]


GVCF_BLOCK_COLUMNS = ["层级路径", "顺序", "字段", "类型/编码", "条件", "说明", "来源"]
GVCF_BLOCK_ROWS = [
    ("CompressedGVCFBlock", "1", "block magic", "4 bytes: `G V C F`", "always", "块内自带二级 magic。", "这不是文件顶层 magic，而是 block payload 自己的 header。", "src/gvcf/gvcf_block.cpp::Serialize()/Deserialize()"),
    ("CompressedGVCFBlock", "2", "block version", "uint8_t", "always", "当前固定写 `1`。", "反序列化时若不是 1 会失败。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "3", "variant_count", "VarUint", "always", "块内变体数。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "4", "sample_count", "VarUint", "always", "块内样本数。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "5", "flags", "uint8_t", "always", "bit0=`has_end_field`，bit1=`has_min_dp`。", "控制后续是否存在 `info_end` / `min_dp` / `dp_min_dp_diff`。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "6", "chrom", "CompressedField", "always", "位置字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "7", "pos", "CompressedField", "always", "位置字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "8", "id", "CompressedField", "always", "位置字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "9", "ref", "CompressedField", "always", "序列字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "10", "alt", "CompressedField", "always", "序列字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "11", "qual", "CompressedField", "always", "质量字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "12", "filter", "CompressedField", "always", "质量字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "13", "info_end", "CompressedField", "`flags & 0x01`", "INFO/END。", "仅 `has_end_field=true` 时序列化。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "14", "gt_mask", "CompressedField", "always", "GT dominant mask。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "15", "gt_patches", "CompressedField", "always", "GT 非 dominant patch 数据。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "16", "gt_phase", "CompressedField", "always", "相位信息。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "17", "dp", "CompressedField", "always", "FORMAT/DP。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "18", "gq", "CompressedField", "always", "FORMAT/GQ。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "19", "min_dp", "CompressedField", "`flags & 0x02`", "FORMAT/MIN_DP。", "仅 `has_min_dp=true` 时序列化。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "20", "dp_min_dp_diff", "CompressedField", "`flags & 0x02`", "`DP - MIN_DP` 差分字段。", "仅 `has_min_dp=true` 时序列化。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "21", "pl", "CompressedField", "always", "FORMAT/PL。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "22", "ad", "CompressedField", "always", "FORMAT/AD。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "23", "unknown_info_count", "VarUint", "always", "未知 INFO 字段个数。", "后面重复 `[name_len][name][CompressedField]`。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "24", "unknown_info_item", "`name_len:VarUint + name bytes + CompressedField`", "repeat", "未知 INFO 字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "25", "unknown_format_count", "VarUint", "always", "未知 FORMAT 字段个数。", "后面重复 `[name_len][name][CompressedField]`。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedGVCFBlock", "26", "unknown_format_item", "`name_len:VarUint + name bytes + CompressedField`", "repeat", "未知 FORMAT 字段。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedField wrapper", "27", "method", "uint8_t", "always", "`FieldCompressionMethod` 枚举值。", "告诉 reader 如何解释 `data`。", "src/gvcf/gvcf_block.cpp"),
    ("CompressedField wrapper", "28", "original_count", "VarUint", "always", "原始元素个数。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedField wrapper", "29", "data_size", "VarUint", "always", "payload 大小。", "/", "src/gvcf/gvcf_block.cpp"),
    ("CompressedField wrapper", "30", "data", "bytes", "always", "具体编码器序列化出来的 payload。", "字段真实格式由 `method` 决定。", "src/gvcf/gvcf_block.cpp; src/gvcf/gvcf_encoding.cpp"),
]


GVCF_ENCODING_COLUMNS = ["编码器 / 方法", "on-disk 结构", "说明", "来源"]
GVCF_ENCODING_ROWS = [
    ("VarUint", "7-bit continuation varint", "最低 7 bit 存值，高位 bit 作为续字节标记。", "src/gvcf/gvcf_encoding.cpp::VarIntUtil"),
    ("VarInt", "`ZigzagEncode(int64)` 后再写 VarUint", "有符号差值/整数统一这样落盘。", "src/gvcf/gvcf_encoding.cpp::VarIntUtil"),
    ("RLE(string)", "`[original_len][run_count]{[strlen][str bytes][count]}`", "字符串数组 RLE。", "src/gvcf/gvcf_encoding.cpp::RLEEncoder"),
    ("RLE(int)", "`[original_len][run_count]{[value_varint][count_varuint]}`", "整数数组 RLE。", "src/gvcf/gvcf_encoding.cpp::RLEIntEncoder"),
    ("RLE(byte)", "`[original_len][run_count]{[value:u8][count_varuint]}`", "字节流 / 位图的 RLE。", "src/gvcf/gvcf_encoding.cpp::RLEByteEncoder"),
    ("Delta", "`[original_len][first_value][delta_count][delta_varint...]`", "POS / END / patch indices 常用。", "src/gvcf/gvcf_encoding.cpp::DeltaEncoder"),
    ("Mask(string)", "`[original_len][dominant_len][dominant][mask_size][RLEByte(mask)][patch_count][idx_size?][Delta(indices)][patch strings...]`", "dominant value + patch 模式。", "src/gvcf/gvcf_encoding.cpp::MaskEncoder"),
    ("Mask(int)", "`[original_len][dominant_value][mask_size][RLEByte(mask)][patch_count][idx_size?][Delta(indices)][patch varint...]`", "整数版 dominant + patch。", "src/gvcf/gvcf_encoding.cpp::MaskIntEncoder"),
    ("Dictionary(string)", "`[original_len][dict_size]{[len][bytes]}[use_rle:u8](if 1 -> [rle_size][RLEInt(indices)] else [idx_count][idx...])`", "字典值本体后面跟索引流。", "src/gvcf/gvcf_encoding.cpp::DictEncoder"),
    ("Dictionary(int)", "`[original_len][dict_size]{[value_varint]}[use_rle:u8](if 1 -> [rle_size][RLEInt(indices)] else [idx_count][idx...])`", "整数版字典编码。", "src/gvcf/gvcf_encoding.cpp::DictIntEncoder"),
]


CONSTANT_COLUMNS = ["类别", "名称", "数值", "含义 / 备注", "来源"]
CONSTANT_ROWS = [
    ("多样本 `.gsc`", "GSC_FIXED_FIELDS_RB_MAGIC", "0x46435347", "fixed-fields chunk magic，little-endian 观感为 `GSCF`。", "src/variant.h"),
    ("多样本 `.gsc`", "GSC_FIXED_FIELDS_RB_VERSION_V1", "1", "旧目录格式：GT 在 chunk header 里单独给 offset/size。", "src/variant.h"),
    ("多样本 `.gsc`", "GSC_FIXED_FIELDS_RB_VERSION_V2", "2", "当前目录格式：每个 row block 自带 `gt_off/gt_size`。", "src/variant.h"),
    ("多样本 `.gsc`", "GSC_META_MAGIC", "0x4D435347", "meta 区 magic。", "src/defs.h"),
    ("多样本 `.gsc`", "GT marker 0", "0", "GT payload backend = BSC。", "src/compressor.cpp"),
    ("多样本 `.gsc`", "GT marker 1", "1", "GT payload backend = Zstd。", "src/compressor.cpp"),
    ("多样本 `.gsc`", "GT marker 2", "2", "GT payload backend = Brotli。", "src/compressor.cpp"),
    ("单样本 `GVCF`", "GVCF_FILE_MAGIC", "0x47564346", "文件级 magic，ASCII 为 `GVCF`。", "src/gvcf/gvcf_compressor.h"),
    ("单样本 `GVCF`", "GVCF_FILE_VERSION", "4", "当前 writer 文件版本。", "src/gvcf/gvcf_compressor.h"),
    ("单样本 `GVCF`", "Block flags bit0", "0x01", "`has_end_field`。", "src/gvcf/gvcf_block.cpp"),
    ("单样本 `GVCF`", "Block flags bit1", "0x02", "`has_min_dp`。", "src/gvcf/gvcf_block.cpp"),
    ("单样本 `GVCF`", "FieldCompressionMethod::NONE", "0", "无压缩或原样。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::RLE", "1", "RLE。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::DELTA", "2", "Delta。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::MASK", "3", "dominant + patch。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::DICTIONARY", "4", "字典编码。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::DELTA_RLE", "5", "Delta + RLE。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::MASK_RLE", "6", "Mask with RLE-compressed bitmask。", "src/gvcf/gvcf_block.h"),
    ("单样本 `GVCF`", "FieldCompressionMethod::ADAPTIVE", "7", "编码器内部自适应选择；通常在真正写盘前会落成具体方法。", "src/gvcf/gvcf_block.h; src/gvcf/gvcf_field_compress.cpp"),
]


SHEETS = [
    ("说明", INFO_COLUMNS, INFO_ROWS),
    ("总览结构", OVERVIEW_COLUMNS, OVERVIEW_ROWS),
    ("多样本主格式", MULTI_ARCHIVE_COLUMNS, MULTI_ARCHIVE_ROWS),
    ("多样本FixedFields", FIXED_FIELDS_COLUMNS, FIXED_FIELDS_ROWS),
    ("多样本Part2", PART2_COLUMNS, PART2_ROWS),
    ("多样本GT补充", GT_COLUMNS, GT_ROWS),
    ("Part2_FMT特殊编码", FMT_SPECIAL_COLUMNS, FMT_SPECIAL_ROWS),
    ("gVCF文件", GVCF_FILE_COLUMNS, GVCF_FILE_ROWS),
    ("gVCF块字段", GVCF_BLOCK_COLUMNS, GVCF_BLOCK_ROWS),
    ("gVCF基础编码", GVCF_ENCODING_COLUMNS, GVCF_ENCODING_ROWS),
    ("常量枚举", CONSTANT_COLUMNS, CONSTANT_ROWS),
]


def srow(*vals: str | None) -> tuple[str | None, ...]:
    padded = list(vals[:12])
    padded.extend([None] * (12 - len(padded)))
    return tuple(padded)


OVERVIEW_WB_INFO_ROWS = [
    ("工作簿定位", "这是“总体结构表”风格的 Excel，主要模仿 `docs/0_avsStructure/GSC多样本二进制格式结构表-260413.xlsx` 的浏览方式。"),
    ("覆盖范围", "同时覆盖多样本 `.gsc` 与单样本 `GVCF` native 文件两条 on-disk 格式主线。"),
    ("与详细工作簿关系", "本工作簿强调层次树和总体结构；字段级、兼容分支和编码细节请配合 `GSC_binary_format_260423.xlsx` 一起看。"),
    ("多样本主线", "参考 `src/compressor.cpp::writeCompressFlie()`、`src/decompression_reader.cpp::OpenReading()/readFixedFields()`、`src/file_handle.cpp`。"),
    ("gVCF 主线", "参考 `src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/CompressAndWriteBlock()/WriteFileFooter()`、`src/gvcf/gvcf_block.cpp`。"),
    ("当前 writer 版本", "多样本 fixed-fields chunk 当前是 `GSCF V2`；gVCF 文件当前是 `GVCF_FILE_VERSION = 4`。"),
]

STRUCTURE_WIDTHS = [20, 10, 18, 20, 10, 18, 22, 10, 22, 22, 10, 22]
STRUCTURE_HEADER_ROW = [
    "Primary Structure", None, None,
    "Secondary Structure", None, None,
    "Tertiary Structure", None, None,
    "Quaternary Structure", None, None,
]
STRUCTURE_LABEL_ROW = [
    "元素名称(序号)\n中文描述", "是否必选", "元素数据示例",
    "元素名称(序号)\n中文描述", "是否必选", "元素数据示例",
    "元素名称(序号)\n中文描述", "是否必选", "元素数据示例",
    "元素名称(序号)\n中文描述", "是否必选", "元素数据示例",
]

STRUCT_ROWS_OVERVIEW = [
    srow("/\nGSC / GVCF 二进制格式", "Y", "[当前源码实际格式]\n多样本 .gsc + 单样本 GVCF"),
    srow("MULTI_SAMPLE_GSC(1)\n多样本 .gsc", "Y", "↘"),
    srow(None, None, None, "TOP_HEADER(1)\n顶层头部", "Y", "↘"),
    srow(None, None, None, None, None, None, "MODE_TYPE / OFFSETS(1-3)\n模式与区段偏移", "Y", "[bool] + [uint64] + [uint64]"),
    srow(None, None, None, "MAIN_ARCHIVE(2)\n主 archive", "Y", "↘"),
    srow(None, None, None, None, None, None, "CHUNK_STREAMS / META / CHUNKS(1)\n定位表、meta 与 chunk 载荷", "Y", "[顺序布局]"),
    srow(None, None, None, None, None, None, "LOSSLESS_PART2(2)\nother fields 多流容器", "lossless", "[File_Handle_2]"),
    srow(None, None, None, None, None, None, "SDSL_TAIL(3)\nzero/copy 位图尾区", "Y", "[4 x rrr_vector]"),
    srow("SINGLE_SAMPLE_GVCF(2)\n单样本 GVCF 文件", "Y", "↘"),
    srow(None, None, None, "FILE_HEADER(1)\n文件头", "Y", "↘"),
    srow(None, None, None, None, None, None, "MAGIC / VERSION / BACKEND(1)\n文件识别与版本", "Y", "[uint32][uint32][uint8]"),
    srow(None, None, None, None, None, None, "HEADER_TEXT / SAMPLE_NAMES(2)\n头文本与样本名", "Y", "↘"),
    srow(None, None, None, "BLOCK_AREA(2)\n块区", "Y", "↘"),
    srow(None, None, None, None, None, None, "BLOCK_ENTRY(1)\n单个块", "Y * N", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "BLOCK_SIZE(1)\n块长度前缀", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, None, None, None, "BLOCK_PAYLOAD(2)\nCompressedGVCFBlock", "Y", "[bytes]"),
    srow(None, None, None, "FILE_FOOTER(3)\n文件尾", "Y", "↘"),
    srow(None, None, None, None, None, None, "BLOCK_OFFSETS / BLOCK_INDEX(1)\n块偏移表与范围索引", "Y", "↘"),
    srow(None, None, None, None, None, None, "FOOTER_OFFSET(2)\n文件尾指针", "Y", "[uint64 @ EOF]"),
]

STRUCT_ROWS_MULTI = [
    srow("/\n.gsc 文件", "Y", "[当前实现]\n无顶层 magic/version"),
    srow("TOP_HEADER(1)\n顶层头部", "Y", "↘"),
    srow(None, None, None, "MODE_TYPE(1)\n文件模式", "Y", "[bool]\n0: lossy\n1: lossless"),
    srow(None, None, None, "OTHER_FIELDS_OFFSET(2)\nother fields 起始偏移", "Y", "[uint64]"),
    srow(None, None, None, "SDSL_OFFSET(3)\nSDSL 尾区起始偏移", "Y", "[uint64]"),
    srow("MAIN_ARCHIVE(2)\n主 archive", "Y", "↘"),
    srow(None, None, None, "CHUNK_STREAMS(1)\nchunk 定位表", "Y", "↘"),
    srow(None, None, None, None, None, None, "CHUNK_STREAMS_SIZE(1)\n边界项数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "CHUNK_STREAM_ITEM(2)\n单个边界项", "Y * N", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "CUR_CHUNK_ACTUAL_POS(1)\n累计真实变体数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, None, None, None, "CHUNK_FILE_OFFSET(2)\nchunk 文件绝对偏移", "Y", "[size_t]"),
    srow(None, None, None, "BASIC_PARAMS(2)\n基础参数", "Y", "↘"),
    srow(None, None, None, None, None, None, "PLOIDY / MAX_BLOCK_ROWS / MAX_BLOCK_COLS(1)\n块参数", "Y", "[uint8] + [uint32]"),
    srow(None, None, None, None, None, None, "VEC_LEN / NO_VEC / COPY_NO(2)\nGT 向量统计", "Y", "[uint64]"),
    srow(None, None, None, None, None, None, "COPY_ORIGIN_BITMAP(3)\ncopy-origin 位级映射", "Y", "[used_bits_cp][size][bytes]"),
    srow(None, None, None, None, None, None, "N_SAMPLES / CHUNKS_MIN_POS / WHERE_CHROM(4)\n样本数与范围索引", "Y", "↘"),
    srow(None, None, None, "COLUMN_TILING(3)\nGT 列分块元数据", "Y", "↘"),
    srow(None, None, None, None, None, None, "N_COL_BLOCKS(1)\n列块数量", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "COL_BLOCK_RANGE_ITEM(2)\n列块范围项", "Y * N", "[start_haplotype][block_size]"),
    srow(None, None, None, "PERMUTATIONS(4)\n排列信息", "Y", "↘"),
    srow(None, None, None, None, None, None, "PERMUTATION_COUNT(1)\n排列条目数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "PERMUTATION_ENTRY(new)(2)\n(row_block,col_block,data)", "conditional", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "PERMUTATION_ENTRY(legacy)(3)\n(row_block,data)", "compat", "↘"),
    srow(None, None, None, "META(5)\nheader / sample 元数据", "Y", "↘"),
    srow(None, None, None, None, None, None, "META_MAGIC / META_BACKEND(1)\nmeta codec 标记", "Y", "[uint32] + [uint8]"),
    srow(None, None, None, None, None, None, "COMP_V_HEADER(2)\n压缩 header 文本", "Y", "[uint32 size] + [bytes]"),
    srow(None, None, None, None, None, None, "COMP_V_SAMPLES(3)\n压缩 sample 列表", "Y", "[uint32 size] + [bytes]"),
    srow(None, None, None, "FIXED_FIELDS_CHUNKS(6)\n所有 chunk 载荷", "Y", "[bytes concat]\n依赖 chunks_streams.offset"),
    srow("LOSSLESS_PART2(3)\nother fields 容器", "lossless", "↘"),
    srow(None, None, None, "PAYLOAD_AREA(1)\n各 stream part 本体", "Y", "[bytes]"),
    srow(None, None, None, "SPECIAL_STREAMS(2)\n特殊命名流", "Y", "↘"),
    srow(None, None, None, None, None, None, "PART2_PARAMS(1)\nactual_variants + keys + dicts", "Y", "[backend prefix] + [compressed payload]"),
    srow(None, None, None, None, None, None, "KEY_i_SIZE / KEY_i_DATA(2)\n每个字段的 size/data 子流", "Y", "[stream_name = key_<i>_*]"),
    srow(None, None, None, "FOOTER(3)\nstream 目录尾部", "Y", "↘"),
    srow(None, None, None, None, None, None, "N_STREAMS / STREAM_ENTRY(1)\n目录项", "Y", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "FOOTER_SIZE(2)\n尾部长度", "Y", "[fixed 8 bytes]"),
    srow("SDSL_TAIL(4)\nGT zero/copy 尾区", "Y", "↘"),
    srow(None, None, None, "ZEROS_ONLY(1)\n全零向量位图", "Y", "[rrr_vector parity 0/1]"),
    srow(None, None, None, "COPY_BIT_VECTOR(2)\ncopy 向量位图", "Y", "[rrr_vector parity 0/1]"),
]

STRUCT_ROWS_FIXED = [
    srow("/\nGSC fixed-fields chunk", "Y", "[magic = GSCF]"),
    srow("GSC_FIXED_FIELDS_CHUNK(1)\n当前 writer 输出格式", "Y", "↘"),
    srow(None, None, None, "CHUNK_HEADER(1)\nchunk 头部", "Y", "↘"),
    srow(None, None, None, None, None, None, "MAGIC(1)\n固定为 GSCF", "Y", "[uint32]\n0x46435347"),
    srow(None, None, None, None, None, None, "VERSION(2)\n目录版本", "Y", "[uint32]\n2"),
    srow(None, None, None, None, None, None, "TOTAL_VARIANTS(3)\nchunk 总变体数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "ROW_BLOCK_COUNT(4)\nrow block 数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "FLAGS(5)\n保留标记", "Y", "[uint32]\n0"),
    srow(None, None, None, "ROW_BLOCK_DIRECTORY(2)\nrow block 目录", "Y", "↘"),
    srow(None, None, None, None, None, None, "ROW_BLOCK_ENTRY(1)\n单个目录项", "Y * ROW_BLOCK_COUNT", "↘"),
    srow(None, None, None, None, None, None, None, None, None, "VARIANT_COUNT / FIRST_POS / LAST_POS(1-3)\n该行块范围", "Y", "[uint32] + [int64]"),
    srow(None, None, None, None, None, None, None, None, None, "FIELD_OFFSETS(4-10)\nchrom/pos/id/ref/alt/qual/gt", "Y", "[off:uint32][size:uint32]"),
    srow(None, None, None, "DATA_REGION(3)\n数据区", "Y", "↘"),
    srow(None, None, None, None, None, None, "ROW_BLOCK_PAYLOAD_ORDER(1)\n按 row block 顺序写", "Y", "chrom -> pos -> id -> ref -> alt -> qual -> gt"),
    srow(None, None, None, None, None, None, "FIELD_PAYLOAD_SEMANTICS(2)\nfixed fields 与 GT 语义", "Y", "字符串 / 差分位置 / GT index"),
    srow(None, None, None, "LEGACY_COMPAT(4)\n旧格式兼容读取", "compat", "↘"),
    srow(None, None, None, None, None, None, "LEGACY_NO_VARIANTS(1)\n旧块起始字段", "compat", "[uint32]"),
    srow(None, None, None, None, None, None, "LEGACY_FIELD_ENTRY(2)\n旧式 size + bytes", "compat", "[uint32 size] + [payload]"),
]

STRUCT_ROWS_GVCF = [
    srow("/\nGVCF native 文件", "Y", "[magic = GVCF]"),
    srow("FILE_HEADER(1)\n文件头", "Y", "↘"),
    srow(None, None, None, "MAGIC(1)\n文件魔数", "Y", "[uint32]\n0x47564346"),
    srow(None, None, None, "VERSION(2)\n文件版本", "Y", "[uint32]\n当前 writer=4"),
    srow(None, None, None, "BACKEND_ID(3)\n压缩后端", "Y", "[uint8]\nV2+"),
    srow(None, None, None, "NUM_SAMPLES(4)\n样本数", "Y", "[uint32]"),
    srow(None, None, None, "HEADER_SECTION(5)\n头文本区", "Y", "↘"),
    srow(None, None, None, None, None, None, "HEADER_FLAG(1)\n头文本是否压缩", "V3+", "[uint8]\n0 raw / 1 compressed"),
    srow(None, None, None, None, None, None, "HEADER_PAYLOAD(2)\n头文本载荷", "Y", "[size fields] + [bytes]"),
    srow(None, None, None, "SAMPLE_NAMES(6)\n样本名列表", "Y", "[repeat name_size + name]"),
    srow(None, None, None, "PLACEHOLDER(7)\n头部尾占位", "Y", "[uint64]"),
    srow("BLOCK_AREA(2)\n块区", "Y", "↘"),
    srow(None, None, None, "BLOCK_ENTRY(1)\n单个压缩块", "Y * N", "↘"),
    srow(None, None, None, None, None, None, "BLOCK_SIZE(1)\n块字节数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "BLOCK_PAYLOAD(2)\nCompressedGVCFBlock", "Y", "[bytes]"),
    srow("FILE_FOOTER(3)\n文件尾", "Y", "↘"),
    srow(None, None, None, "NUM_BLOCKS(1)\n块数量", "Y", "[uint32]"),
    srow(None, None, None, "BLOCK_OFFSETS(2)\n块偏移表", "Y", "[uint64] * N"),
    srow(None, None, None, "TOTAL_VARIANTS(3)\n总变体数", "Y", "[uint64]"),
    srow(None, None, None, "BLOCK_INDEX_AREA(4)\n范围查询索引", "V4", "↘"),
    srow(None, None, None, None, None, None, "NUM_INDICES(1)\n索引个数", "Y", "[uint32]"),
    srow(None, None, None, None, None, None, "BLOCK_INDEX_ITEM(2)\n单个块索引", "Y * N", "chrom_len + chrom + start/end + offset + count"),
    srow(None, None, None, "FOOTER_OFFSET(5)\n文件尾指针", "Y", "[uint64 @ EOF]"),
]

STRUCT_ROWS_GVCF_BLOCK = [
    srow("/\nCompressedGVCFBlock", "Y", "[block payload body]"),
    srow("BLOCK_HEADER(1)\n块内部头", "Y", "↘"),
    srow(None, None, None, "MAGIC(1)\n块 magic", "Y", "[bytes]\nGVCF"),
    srow(None, None, None, "BLOCK_VERSION(2)\n块版本", "Y", "[uint8]\n1"),
    srow(None, None, None, "VARIANT_COUNT / SAMPLE_COUNT(3-4)\n块级统计", "Y", "[VarUint]"),
    srow(None, None, None, "FLAGS(5)\n字段存在位", "Y", "[uint8]\nbit0 END / bit1 MIN_DP"),
    srow("KNOWN_FIELDS(2)\n已知字段", "Y", "↘"),
    srow(None, None, None, "POSITION_FIELDS(1)\n位置字段", "Y", "↘"),
    srow(None, None, None, None, None, None, "CHROM / POS / ID(1-3)\n位置三元组", "Y", "[CompressedField]"),
    srow(None, None, None, "SEQUENCE_FIELDS(2)\n序列字段", "Y", "↘"),
    srow(None, None, None, None, None, None, "REF / ALT(1-2)\n参考与替换", "Y", "[CompressedField]"),
    srow(None, None, None, "QUALITY_FIELDS(3)\n质量字段", "Y", "↘"),
    srow(None, None, None, None, None, None, "QUAL / FILTER(1-2)\n质量与过滤", "Y", "[CompressedField]"),
    srow(None, None, None, "INFO_FIELDS(4)\nINFO 字段", "conditional", "↘"),
    srow(None, None, None, None, None, None, "INFO_END(1)\nEND 范围字段", "has_end_field", "[CompressedField]"),
    srow(None, None, None, "SAMPLE_FIELDS(5)\n样本字段", "Y", "↘"),
    srow(None, None, None, None, None, None, "GT_MASK / GT_PATCHES / GT_PHASE(1-3)\nGT 相关", "Y", "[CompressedField]"),
    srow(None, None, None, None, None, None, "DP / GQ / MIN_DP / DP_MIN_DP_DIFF(4-7)\n深度与质量", "conditional", "[CompressedField]"),
    srow(None, None, None, None, None, None, "PL / AD(8-9)\n样本 likelihood/depth", "Y", "[CompressedField]"),
    srow("UNKNOWN_INFO(3)\n未知 INFO", "Y", "↘"),
    srow(None, None, None, "COUNT + ITEM(1)\n未知 INFO 列表", "Y", "[name_len][name][CompressedField]"),
    srow("UNKNOWN_FORMAT(4)\n未知 FORMAT", "Y", "↘"),
    srow(None, None, None, "COUNT + ITEM(1)\n未知 FORMAT 列表", "Y", "[name_len][name][CompressedField]"),
    srow("COMPRESSED_FIELD_WRAPPER(5)\n字段统一包装", "Y", "↘"),
    srow(None, None, None, "METHOD(1)\n压缩方法枚举", "Y", "[uint8]"),
    srow(None, None, None, "ORIGINAL_COUNT(2)\n原始元素个数", "Y", "[VarUint]"),
    srow(None, None, None, "DATA_SIZE(3)\npayload 大小", "Y", "[VarUint]"),
    srow(None, None, None, "DATA(4)\n编码器序列化 bytes", "Y", "[bytes]"),
]


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
WRAP_ALIGN = Alignment(vertical="top", horizontal="left", wrap_text=True)
STRUCTURE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="CFE2F3")
STRUCTURE_SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="EAF3FB")


def write_sheet(wb: Workbook, title: str, columns: list[str], rows: list[tuple[str, ...]]) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    for c_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = WRAP_ALIGN

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in rows:
            value = row[idx - 1]
            if value is None:
                continue
            for part in str(value).splitlines():
                max_len = max(max_len, len(part))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 64)


def write_structure_sheet(wb: Workbook, title: str, rows: list[tuple[str | None, ...]]) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A3"

    for rng in ("A1:C1", "D1:F1", "G1:I1", "J1:L1"):
        ws.merge_cells(rng)

    for idx, value in enumerate(STRUCTURE_HEADER_ROW, start=1):
        cell = ws.cell(row=1, column=idx, value=value)
        cell.fill = STRUCTURE_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    for idx, value in enumerate(STRUCTURE_LABEL_ROW, start=1):
        cell = ws.cell(row=2, column=idx, value=value)
        cell.fill = STRUCTURE_SUBHEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    for r_idx, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = WRAP_ALIGN

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42
    for idx, width in enumerate(STRUCTURE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def md_escape(value: str) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def render_table(columns: list[str], rows: list[tuple[str, ...]]) -> str:
    header = "| " + " | ".join(columns) + " |"
    sep = "| " + " | ".join("---" for _ in columns) + " |"
    body = []
    for row in rows:
        body.append("| " + " | ".join(md_escape(x) for x in row) + " |")
    return "\n".join([header, sep, *body])


def write_markdown() -> None:
    chunks: list[str] = []
    chunks.append("# GSC 二进制格式总整理（260423）")
    chunks.append("")
    chunks.append("> 目录：`docs/0423_gsc_binary_format/`")
    chunks.append("")
    chunks.append("本文把当前仓库里的两条实际二进制主线统一整理为一份总文档：")
    chunks.append("")
    chunks.append("- 多样本 VCF/BCF 压缩输出的 `.gsc`")
    chunks.append("- 单样本 `gsc gvcf` 输出的 `GVCF` native 文件")
    chunks.append("")
    chunks.append("参考了 `docs/0_avsStructure/` 的现有材料，但所有字段顺序、类型宽度、条件分支和兼容逻辑均以当前源码 writer/reader 为准。")
    chunks.append("")
    chunks.append("## 1. 说明")
    chunks.append("")
    chunks.append(render_table(INFO_COLUMNS, INFO_ROWS))
    chunks.append("")
    chunks.append("## 2. 顶层总览")
    chunks.append("")
    chunks.append("多样本 `.gsc` 当前是三段主体加一个顶层头：")
    chunks.append("")
    chunks.append("```text")
    chunks.append("[top header]")
    chunks.append("[main archive]")
    chunks.append("[lossless part2 region]   // only when mode_type=true")
    chunks.append("[sdsl tail]")
    chunks.append("```")
    chunks.append("")
    chunks.append("单样本 `GVCF` 当前是：")
    chunks.append("")
    chunks.append("```text")
    chunks.append("[file header]")
    chunks.append("[block_size + block_payload] * N")
    chunks.append("[file footer]")
    chunks.append("[footer_offset at EOF]")
    chunks.append("```")
    chunks.append("")
    chunks.append(render_table(OVERVIEW_COLUMNS, OVERVIEW_ROWS))
    chunks.append("")
    chunks.append("## 3. 多样本 `.gsc` 主格式")
    chunks.append("")
    chunks.append(render_table(MULTI_ARCHIVE_COLUMNS, MULTI_ARCHIVE_ROWS))
    chunks.append("")
    chunks.append("## 4. `GSCF` Fixed-Fields Chunk")
    chunks.append("")
    chunks.append(render_table(FIXED_FIELDS_COLUMNS, FIXED_FIELDS_ROWS))
    chunks.append("")
    chunks.append("## 5. Lossless Part2 容器")
    chunks.append("")
    chunks.append(render_table(PART2_COLUMNS, PART2_ROWS))
    chunks.append("")
    chunks.append("### 5.1 GT 与全局辅助索引补充")
    chunks.append("")
    chunks.append(render_table(GT_COLUMNS, GT_ROWS))
    chunks.append("")
    chunks.append("### 5.2 Part2 特殊 FORMAT 编码补充")
    chunks.append("")
    chunks.append(render_table(FMT_SPECIAL_COLUMNS, FMT_SPECIAL_ROWS))
    chunks.append("")
    chunks.append("## 6. 单样本 `GVCF` 文件格式")
    chunks.append("")
    chunks.append(render_table(GVCF_FILE_COLUMNS, GVCF_FILE_ROWS))
    chunks.append("")
    chunks.append("## 7. `CompressedGVCFBlock` 内部布局")
    chunks.append("")
    chunks.append(render_table(GVCF_BLOCK_COLUMNS, GVCF_BLOCK_ROWS))
    chunks.append("")
    chunks.append("## 8. gVCF 基础编码器的 payload 形状")
    chunks.append("")
    chunks.append(render_table(GVCF_ENCODING_COLUMNS, GVCF_ENCODING_ROWS))
    chunks.append("")
    chunks.append("## 9. 常量与枚举")
    chunks.append("")
    chunks.append(render_table(CONSTANT_COLUMNS, CONSTANT_ROWS))
    chunks.append("")
    chunks.append("## 10. 结论")
    chunks.append("")
    chunks.append(dedent(
        """
        当前仓库里其实并存两套文件级协议：

        - 多样本 `.gsc`：没有顶层 magic/version，重心在 `main archive + GSCF V2 chunk + part2 + sdsl tail`
        - 单样本 `GVCF`：有明确的 `GVCF` magic/version，采用 `header + blocks + footer_offset` 结构

        真正做兼容、转码或协议对接时，建议优先注意下面几个点：

        - 多样本 `.gsc` 依赖 `bool`、`size_t` 和 `footer_size` 的 8-byte 原样写盘，不是严格跨平台协议。
        - final `.gsc` 里的 chunk 没有长度前缀，必须依赖 `chunks_streams[].offset`。
        - fixed-fields、meta、part2_params 里有不少“数字其实是十进制 ASCII + `\\0`”的 token 流，不能误按固定宽度整数读取。
        - gVCF 文件级格式和多样本 `.gsc` 完全不同，不能共用 reader 假设。
        """
    ).strip())
    chunks.append("")
    OUT_MD.write_text("\n".join(chunks), encoding="utf-8")


def write_workbook() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for title, columns, rows in SHEETS:
        write_sheet(wb, title, columns, rows)

    wb.save(OUT_XLSX)


def write_overview_workbook() -> None:
    wb = Workbook()
    default = wb.active
    default.title = "说明"

    for c_idx, col_name in enumerate(["对象", "说明"], start=1):
        cell = default.cell(row=1, column=c_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    for r_idx, row in enumerate(OVERVIEW_WB_INFO_ROWS, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = default.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = WRAP_ALIGN

    default.column_dimensions["A"].width = 18
    default.column_dimensions["B"].width = 84
    default.freeze_panes = "A2"

    write_structure_sheet(wb, "总体结构", STRUCT_ROWS_OVERVIEW)
    write_structure_sheet(wb, "多样本结构", STRUCT_ROWS_MULTI)
    write_structure_sheet(wb, "多样本FixedFields结构", STRUCT_ROWS_FIXED)
    write_structure_sheet(wb, "gVCF结构", STRUCT_ROWS_GVCF)
    write_structure_sheet(wb, "gVCF块结构", STRUCT_ROWS_GVCF_BLOCK)

    wb.save(OUT_OVERVIEW_XLSX)


def write_gsc_logic_svg() -> None:
    svg = dedent(
        """
        <?xml version="1.0" encoding="UTF-8"?>
        <svg xmlns="http://www.w3.org/2000/svg" width="1400" height="900" viewBox="0 0 1400 900" font-family="Arial, Helvetica, sans-serif">
          <defs>
            <linearGradient id="titleGrad" x1="0%" y1="0%" x2="100%" y2="0%">
              <stop offset="0%" stop-color="#17324D"/>
              <stop offset="100%" stop-color="#28597A"/>
            </linearGradient>
            <linearGradient id="rootGrad" x1="0%" y1="0%" x2="0%" y2="100%">
              <stop offset="0%" stop-color="#E8F2FB"/>
              <stop offset="100%" stop-color="#D6E7F7"/>
            </linearGradient>
            <linearGradient id="panelGradA" x1="0%" y1="0%" x2="0%" y2="100%">
              <stop offset="0%" stop-color="#EEF7F1"/>
              <stop offset="100%" stop-color="#DFF1E5"/>
            </linearGradient>
            <linearGradient id="panelGradB" x1="0%" y1="0%" x2="0%" y2="100%">
              <stop offset="0%" stop-color="#F4F6FB"/>
              <stop offset="100%" stop-color="#E8EEFA"/>
            </linearGradient>
            <linearGradient id="panelGradC" x1="0%" y1="0%" x2="0%" y2="100%">
              <stop offset="0%" stop-color="#FFF6E8"/>
              <stop offset="100%" stop-color="#FDE8C2"/>
            </linearGradient>
            <linearGradient id="panelGradD" x1="0%" y1="0%" x2="0%" y2="100%">
              <stop offset="0%" stop-color="#F7EEF8"/>
              <stop offset="100%" stop-color="#EEDFF2"/>
            </linearGradient>
            <filter id="shadow" x="-8%" y="-8%" width="116%" height="116%">
              <feDropShadow dx="0" dy="5" stdDeviation="7" flood-color="#10253A" flood-opacity="0.14"/>
            </filter>
            <marker id="kiteArrow" markerWidth="14" markerHeight="14" refX="11" refY="7" orient="auto">
              <path d="M1,1 L12,7 L1,13 L4.3,7 Z" fill="#2E5E7A"/>
            </marker>
            <marker id="smallArrow" markerWidth="12" markerHeight="12" refX="9.5" refY="6" orient="auto">
              <path d="M1,1 L10,6 L1,11 L3.8,6 Z" fill="#4D6D87"/>
            </marker>
            <style>
              .title-main { font-size: 28px; font-weight: 700; fill: #FFFFFF; }
              .title-sub { font-size: 13px; fill: #D8E7F4; }
              .section-title { font-size: 17px; font-weight: 700; fill: #17324D; }
              .card-title { font-size: 16px; font-weight: 700; fill: #17324D; }
              .card-sub { font-size: 12px; fill: #4D6274; }
              .chip-title { font-size: 12px; font-weight: 700; fill: #17324D; }
              .chip-text { font-size: 11px; fill: #3D4D5C; }
              .mono { font-size: 11px; fill: #17324D; font-weight: 700; }
              .note { font-size: 12px; fill: #5B6D7A; }
            </style>
          </defs>

          <rect width="1400" height="900" fill="#F6F9FC"/>
          <rect x="0" y="0" width="1400" height="78" fill="url(#titleGrad)"/>
          <text x="700" y="36" text-anchor="middle" class="title-main">GSC二进制封装格式</text>
          <text x="700" y="59" text-anchor="middle" class="title-sub">基于当前源码 writer / reader 的 .gsc on-disk 结构梳理</text>

          <g transform="translate(42,112)" filter="url(#shadow)">
            <rect x="0" y="0" width="248" height="666" rx="22" fill="url(#rootGrad)" stroke="#8FB0CC" stroke-width="1.6"/>
            <text x="124" y="48" text-anchor="middle" class="section-title">`.gsc` 文件整体</text>
            <rect x="24" y="78" width="200" height="110" rx="18" fill="#FFFFFF" stroke="#A8C1D8"/>
            <text x="124" y="113" text-anchor="middle" class="card-title">多样本主线封装</text>
            <text x="124" y="136" text-anchor="middle" class="card-sub">top header + main archive</text>
            <text x="124" y="156" text-anchor="middle" class="card-sub">lossless part2 + sdsl tail</text>

            <rect x="24" y="214" width="200" height="124" rx="18" fill="#FFFDF8" stroke="#E5C676"/>
            <text x="124" y="246" text-anchor="middle" class="card-title">顶层约束</text>
            <text x="124" y="271" text-anchor="middle" class="chip-text">无顶层全局魔数 / version</text>
            <text x="124" y="293" text-anchor="middle" class="chip-text">文件头固定先写 17 字节</text>
            <text x="124" y="315" text-anchor="middle" class="chip-text">lossless 与 lossy 路径分叉</text>

            <rect x="24" y="364" width="200" height="170" rx="18" fill="#FFFFFF" stroke="#A8C1D8"/>
            <text x="124" y="396" text-anchor="middle" class="card-title">关键判断点</text>
            <text x="124" y="422" text-anchor="middle" class="chip-text">1. `mode_type`</text>
            <text x="124" y="446" text-anchor="middle" class="chip-text">2. `other_fields_offset`</text>
            <text x="124" y="470" text-anchor="middle" class="chip-text">3. `sdsl_offset`</text>
            <text x="124" y="500" text-anchor="middle" class="chip-text">fixed-fields chunk 用</text>
            <text x="124" y="522" text-anchor="middle" class="mono">`GSCF` 魔数 + version</text>

            <rect x="24" y="560" width="200" height="84" rx="18" fill="#F1F6FC" stroke="#A8C1D8"/>
            <text x="124" y="593" text-anchor="middle" class="card-title">协议边界</text>
            <text x="124" y="618" text-anchor="middle" class="chip-text">本图只描述 GSC</text>
            <text x="124" y="638" text-anchor="middle" class="chip-text">不再混排 GVCF 文件结构</text>
          </g>

          <g transform="translate(336,118)">
            <rect x="0" y="0" width="264" height="106" rx="20" fill="url(#panelGradA)" stroke="#77B38A" stroke-width="1.8" filter="url(#shadow)"/>
            <text x="26" y="38" class="card-title">TOP_HEADER</text>
            <text x="26" y="62" class="card-sub">顶层头部，先于 archive 写入</text>
            <text x="26" y="84" class="chip-text">`mode_type` / `other_fields_offset` / `sdsl_offset`</text>
          </g>

          <g transform="translate(336,266)">
            <rect x="0" y="0" width="264" height="118" rx="20" fill="url(#panelGradB)" stroke="#8BA6D4" stroke-width="1.8" filter="url(#shadow)"/>
            <text x="26" y="40" class="card-title">MAIN_ARCHIVE</text>
            <text x="26" y="64" class="card-sub">主 archive，承载 GT 索引与 fixed-fields chunk</text>
            <text x="26" y="88" class="chip-text">`chunks_streams` / `meta` / `fixed-fields chunks`</text>
            <text x="26" y="108" class="chip-text">列分块、排列表、copy-origin 映射都在这里</text>
          </g>

          <g transform="translate(336,434)">
            <rect x="0" y="0" width="264" height="106" rx="20" fill="url(#panelGradC)" stroke="#DAAB43" stroke-width="1.8" filter="url(#shadow)"/>
            <text x="26" y="38" class="card-title">LOSSLESS_PART2</text>
            <text x="26" y="62" class="card-sub">only when `mode_type=true`</text>
            <text x="26" y="84" class="chip-text">other fields 多流容器 + footer 目录</text>
          </g>

          <g transform="translate(336,582)">
            <rect x="0" y="0" width="264" height="106" rx="20" fill="url(#panelGradD)" stroke="#B784C4" stroke-width="1.8" filter="url(#shadow)"/>
            <text x="26" y="38" class="card-title">SDSL_TAIL</text>
            <text x="26" y="62" class="card-sub">GT 全局 zero / copy 位图尾区</text>
            <text x="26" y="84" class="chip-text">从 `sdsl_offset` 开始顺序读取四个 `rrr_vector`</text>
          </g>

          <path d="M 290 224 C 322 224, 314 170, 336 170" fill="none" stroke="#2E5E7A" stroke-width="4.2" stroke-linecap="round" marker-end="url(#kiteArrow)"/>
          <path d="M 290 360 C 322 360, 314 325, 336 325" fill="none" stroke="#2E5E7A" stroke-width="4.2" stroke-linecap="round" marker-end="url(#kiteArrow)"/>
          <path d="M 290 506 C 322 506, 314 487, 336 487" fill="none" stroke="#2E5E7A" stroke-width="4.2" stroke-linecap="round" marker-end="url(#kiteArrow)"/>
          <path d="M 290 676 C 322 676, 314 635, 336 635" fill="none" stroke="#2E5E7A" stroke-width="4.2" stroke-linecap="round" marker-end="url(#kiteArrow)"/>

          <g transform="translate(646,94)" filter="url(#shadow)">
            <rect x="0" y="0" width="706" height="154" rx="24" fill="#FFFFFF" stroke="#9CB7C9" stroke-width="1.5"/>
            <text x="30" y="36" class="section-title">TOP_HEADER 详细字段</text>
            <text x="30" y="60" class="note">没有全局文件魔数；reader 直接从 17 字节顶层头开始判定格式走向。</text>

            <rect x="30" y="82" width="194" height="48" rx="14" fill="#EEF7F1" stroke="#8BB89A"/>
            <text x="48" y="103" class="chip-title">`mode_type`</text>
            <text x="48" y="120" class="chip-text">`true=lossless`, `false=lossy`</text>

            <rect x="256" y="82" width="206" height="48" rx="14" fill="#F4F6FB" stroke="#A2B6D8"/>
            <text x="274" y="103" class="chip-title">`other_fields_offset`</text>
            <text x="274" y="120" class="chip-text">lossless part2 的起始偏移</text>

            <rect x="494" y="82" width="182" height="48" rx="14" fill="#F7EEF8" stroke="#C2A4CB"/>
            <text x="512" y="103" class="chip-title">`sdsl_offset`</text>
            <text x="512" y="120" class="chip-text">SDSL 尾区起始偏移</text>

            <rect x="688" y="82" width="18" height="48" rx="9" fill="#D8E6F2"/>
          </g>

          <g transform="translate(646,276)" filter="url(#shadow)">
            <rect x="0" y="0" width="706" height="286" rx="24" fill="#FFFFFF" stroke="#9CB7C9" stroke-width="1.5"/>
            <text x="30" y="38" class="section-title">MAIN_ARCHIVE 内部结构</text>
            <text x="30" y="62" class="note">核心布局来自 `writeCompressFlie()`，解压侧依靠 offsets、meta 和 chunk 边界表恢复定位。</text>

            <rect x="30" y="88" width="150" height="56" rx="16" fill="#F4F6FB" stroke="#A2B6D8"/>
            <text x="49" y="111" class="chip-title">`chunks_streams`</text>
            <text x="49" y="129" class="chip-text">chunk 边界表</text>

            <rect x="198" y="88" width="158" height="56" rx="16" fill="#F4F6FB" stroke="#A2B6D8"/>
            <text x="217" y="111" class="chip-title">basic params</text>
            <text x="217" y="129" class="chip-text">ploidy / vec / samples</text>

            <rect x="374" y="88" width="150" height="56" rx="16" fill="#F4F6FB" stroke="#A2B6D8"/>
            <text x="393" y="111" class="chip-title">column tiling</text>
            <text x="393" y="129" class="chip-text">`n_col_blocks` 等</text>

            <rect x="542" y="88" width="134" height="56" rx="16" fill="#F4F6FB" stroke="#A2B6D8"/>
            <text x="561" y="111" class="chip-title">permutations</text>
            <text x="561" y="129" class="chip-text">排列表</text>

            <rect x="30" y="162" width="160" height="56" rx="16" fill="#EEF7F1" stroke="#8BB89A"/>
            <text x="49" y="185" class="chip-title">meta</text>
            <text x="49" y="203" class="chip-text">header / samples 压缩区</text>

            <rect x="214" y="162" width="238" height="56" rx="16" fill="#FFF6E8" stroke="#E1B55C"/>
            <text x="233" y="185" class="chip-title">fixed-fields chunks</text>
            <text x="233" y="203" class="chip-text">最终文件里顺序拼接，不带 chunk-size 前缀</text>

            <path d="M 459 190 C 497 190, 522 190, 550 190" fill="none" stroke="#70859A" stroke-width="3.1" stroke-linecap="round" marker-end="url(#smallArrow)"/>

            <rect x="470" y="158" width="210" height="96" rx="18" fill="#FFFDF8" stroke="#E5C676"/>
            <text x="489" y="181" class="chip-title">固定字段 chunk 细节</text>
            <text x="489" y="201" class="mono">`GSCF` 魔数 + `version=2`</text>
            <text x="489" y="219" class="chip-text">`row_block_dir` + `data_region`</text>
            <text x="489" y="237" class="chip-text">`chrom/pos/id/ref/alt/qual/gt`</text>

            <rect x="30" y="236" width="650" height="28" rx="14" fill="#EDF3F9" stroke="#C7D6E4"/>
            <text x="48" y="255" class="chip-text">GT index payload、copy-origin bitstream、列块范围信息都依赖这个主 archive 统一定位。</text>
          </g>

          <g transform="translate(646,592)" filter="url(#shadow)">
            <rect x="0" y="0" width="432" height="164" rx="24" fill="#FFFFFF" stroke="#D3B16A" stroke-width="1.5"/>
            <text x="30" y="38" class="section-title">LOSSLESS_PART2 细化</text>
            <text x="30" y="61" class="note">只有 `mode_type=true` 时存在，对应 `[other_fields_offset, sdsl_offset)`。</text>

            <rect x="30" y="86" width="170" height="52" rx="16" fill="#FFF6E8" stroke="#E1B55C"/>
            <text x="48" y="109" class="chip-title">`part2_params`</text>
            <text x="48" y="127" class="chip-text">actual_variants / keys / dicts</text>

            <rect x="218" y="86" width="182" height="52" rx="16" fill="#FFF6E8" stroke="#E1B55C"/>
            <text x="236" y="109" class="chip-title">`key_i_size` / `key_i_data`</text>
            <text x="236" y="127" class="chip-text">每个字段拆成 size/data 子流</text>

            <rect x="30" y="148" width="370" height="16" rx="8" fill="#F5E7BE"/>
          </g>

          <g transform="translate(1102,592)" filter="url(#shadow)">
            <rect x="0" y="0" width="250" height="164" rx="24" fill="#FFFFFF" stroke="#C2A4CB" stroke-width="1.5"/>
            <text x="30" y="38" class="section-title">SDSL_TAIL 细化</text>
            <text x="30" y="61" class="note">GT 全局 zero/copy 索引尾区</text>

            <rect x="30" y="84" width="190" height="30" rx="14" fill="#F7EEF8" stroke="#C2A4CB"/>
            <text x="48" y="103" class="chip-text">`zeros_only[0]` / `zeros_only[1]`</text>

            <rect x="30" y="122" width="190" height="30" rx="14" fill="#F7EEF8" stroke="#C2A4CB"/>
            <text x="48" y="141" class="chip-text">`copy_bit_vector[0]` / `[1]`</text>
          </g>

          <path d="M 600 170 C 624 170, 624 170, 646 170" fill="none" stroke="#3F708B" stroke-width="4" stroke-linecap="round" marker-end="url(#kiteArrow)"/>
          <path d="M 600 325 C 624 325, 624 420, 646 420" fill="none" stroke="#3F708B" stroke-width="4" stroke-linecap="round" marker-end="url(#kiteArrow)"/>
          <path d="M 600 487 C 624 487, 624 674, 646 674" fill="none" stroke="#3F708B" stroke-width="4" stroke-linecap="round" marker-end="url(#kiteArrow)"/>
          <path d="M 600 635 C 624 635, 1004 635, 1102 674" fill="none" stroke="#3F708B" stroke-width="4" stroke-linecap="round" marker-end="url(#kiteArrow)"/>

          <rect x="42" y="820" width="1310" height="44" rx="18" fill="#EAF1F7" stroke="#C5D3DE"/>
          <text x="70" y="848" class="note">源码定位: `src/compressor.cpp::writeCompressFlie()` / `src/decompression_reader.cpp::OpenReading()` / `src/file_handle.cpp` / `src/variant.h`</text>
        </svg>
        """
    ).strip() + "\n"
    OUT_GSC_LOGIC_SVG.write_text(svg, encoding="utf-8")


def main() -> None:
    write_markdown()
    write_workbook()
    write_overview_workbook()
    write_gsc_logic_svg()
    print(f"Wrote {OUT_MD}")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_OVERVIEW_XLSX}")
    print(f"Wrote {OUT_GSC_LOGIC_SVG}")


if __name__ == "__main__":
    main()
