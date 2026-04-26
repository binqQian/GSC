from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from generate_vcf_gvcf_fields import ROWS


ROOT = Path(__file__).resolve().parent
OUT_XLSX = ROOT / "VCF_gVCF二进制格式结构表-260424.xlsx"


STRUCTURE_WIDTHS = [22, 10, 22, 24, 10, 24, 28, 10, 30, 34, 14, 52]
STRUCTURE_HEADER_ROW = [
    "Primary Structure",
    None,
    None,
    "Secondary Structure",
    None,
    None,
    "Tertiary Structure",
    None,
    None,
    "Quaternary Structure",
    None,
    None,
]
STRUCTURE_LABEL_ROW = [
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
    "元素名称(序号)\n中文描述",
    "是否必选",
    "元素数据示例",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="CFE2F3")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="EAF3FB")
VCF_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
GVCF_FILL = PatternFill(fill_type="solid", fgColor="EAF3FB")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FONT = Font(bold=True)
WRAP_ALIGN = Alignment(vertical="top", horizontal="left", wrap_text=True)


VCF_SECTION_ORDER = {
    "TOP_HEADER": 10,
    "MAIN_ARCHIVE": 20,
    "GSCF_CHUNK": 30,
    "LOSSLESS_PART2": 40,
    "SDSL_TAIL": 50,
    "VCF_LOGICAL_FIELDS": 60,
    "VCF_FORMAT_SPECIAL": 70,
    "GT_SPARSE_INDEX": 80,
}

GVCF_SECTION_ORDER = {
    "FILE_HEADER": 10,
    "BLOCK_AREA": 20,
    "CompressedGVCFBlock": 30,
    "CompressedField": 40,
    "FILE_FOOTER": 50,
    "ENCODING": 60,
    "FIELD_PAYLOAD": 70,
}

SECTION_LABELS = {
    "TOP_HEADER": "TOP_HEADER\n.gsc 顶层头部",
    "MAIN_ARCHIVE": "MAIN_ARCHIVE\n.gsc 主 archive",
    "GSCF_CHUNK": "GSCF_CHUNK\nfixed fields chunk",
    "LOSSLESS_PART2": "LOSSLESS_PART2\nlossless other fields 容器",
    "SDSL_TAIL": "SDSL_TAIL\nzero/copy 位图尾区",
    "VCF_LOGICAL_FIELDS": "VCF_LOGICAL_FIELDS\nVCF 字段覆盖范围",
    "VCF_FORMAT_SPECIAL": "VCF_FORMAT_SPECIAL\n特殊 FORMAT codec",
    "GT_SPARSE_INDEX": "GT_SPARSE_INDEX\nGT 稀疏索引语义",
    "FILE_HEADER": "FILE_HEADER\nnative GVCF 文件头",
    "BLOCK_AREA": "BLOCK_AREA\nnative GVCF 块区",
    "CompressedGVCFBlock": "COMPRESSED_GVCF_BLOCK\n块 payload 内部结构",
    "CompressedField": "COMPRESSED_FIELD\n字段统一包装",
    "FILE_FOOTER": "FILE_FOOTER\nnative GVCF 文件尾",
    "ENCODING": "ENCODING\ngVCF 基础编码器",
    "FIELD_PAYLOAD": "FIELD_PAYLOAD\ngVCF 字段 payload 形状",
}

EXCLUDED_TOP_LEVELS = {
    "VCF_INTERNAL",  # internal field_desc/key_desc helpers, not standalone on-disk sections
    "GVCFBlock",    # raw pre-compression view; on-disk fields are under CompressedGVCFBlock
}


def srow(*vals: str | None) -> tuple[str | None, ...]:
    padded = list(vals[:12])
    padded.extend([None] * (12 - len(padded)))
    return tuple(padded)


def section_order(fmt: str, top: str) -> int:
    if fmt == "VCF/.gsc":
        return VCF_SECTION_ORDER.get(top, 999)
    return GVCF_SECTION_ORDER.get(top, 999)


def filtered_rows() -> list[tuple[str, ...]]:
    rows: list[tuple[str, ...]] = []
    for row in ROWS:
        fmt, path = row[0], row[1]
        top = path.split("/", 1)[0]
        if top in EXCLUDED_TOP_LEVELS:
            continue
        rows.append(row)
    return sorted(
        rows,
        key=lambda r: (
            0 if r[0] == "VCF/.gsc" else 1,
            section_order(r[0], r[1].split("/", 1)[0]),
            ROWS.index(r),
        ),
    )


def describe_leaf(field: str, semantic: str) -> str:
    if semantic and semantic != field:
        return f"{field}\n{semantic}"
    return field


def describe_payload(encoding: str, order: str, note: str) -> str:
    parts = [encoding]
    if order:
        parts.append(f"顺序/重复: {order}")
    if note:
        parts.append(note)
    return "\n".join(parts)


def make_structure_rows() -> list[tuple[str | None, ...]]:
    rows: list[tuple[str | None, ...]] = [
        srow("/\nVCF / gVCF 当前二进制封装结构", "Y", "仅记录本项目真实写出的字段和顺序\n不补入 AVS 规范中未实现的外层节点"),
    ]

    current_fmt: str | None = None
    current_top: str | None = None
    current_second: tuple[str, str] | None = None
    current_third: tuple[str, str, str] | None = None

    primary_index = {"VCF/.gsc": "1", "gVCF/native": "2"}
    primary_label = {
        "VCF/.gsc": "MULTI_SAMPLE_GSC(1)\n多样本 VCF/BCF .gsc 文件",
        "gVCF/native": "NATIVE_GVCF(2)\n单样本 native GVCF 文件",
    }
    primary_example = {
        "VCF/.gsc": "top header + main archive + optional part2 + sdsl tail",
        "gVCF/native": "file header + block area + footer + footer_offset",
    }

    for fmt, path, field, semantic, encoding, required, order, note, _source in filtered_rows():
        parts = path.split("/")
        top = parts[0]
        second = parts[1] if len(parts) > 1 else ""
        third = parts[2] if len(parts) > 2 else ""

        if fmt != current_fmt:
            rows.append(srow(primary_label[fmt], "Y", primary_example[fmt]))
            current_fmt = fmt
            current_top = None
            current_second = None
            current_third = None

        if top != current_top:
            label = SECTION_LABELS.get(top, top)
            rows.append(srow(None, None, None, f"{label}", "Y", "↘"))
            current_top = top
            current_second = None
            current_third = None

        if not second:
            rows.append(srow(
                None, None, None,
                None, None, None,
                describe_leaf(field, semantic), required, describe_payload(encoding, order, note),
            ))
            continue

        second_key = (top, second)
        if second_key != current_second:
            rows.append(srow(None, None, None, None, None, None, f"{second}", "Y/conditional", "↘"))
            current_second = second_key
            current_third = None

        if not third:
            rows.append(srow(
                None, None, None,
                None, None, None,
                None, None, None,
                describe_leaf(field, semantic), required, describe_payload(encoding, order, note),
            ))
            continue

        third_key = (top, second, third)
        if third_key != current_third:
            rows.append(srow(
                None, None, None,
                None, None, None,
                f"{third}", "Y/conditional", "↘",
            ))
            current_third = third_key

        rows.append(srow(
            None, None, None,
            None, None, None,
            None, None, None,
            describe_leaf(field, semantic), required, describe_payload(encoding, order, note),
        ))

    return rows


def write_structure_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "结构"
    ws.freeze_panes = "A3"

    for rng in ("A1:C1", "D1:F1", "G1:I1", "J1:L1"):
        ws.merge_cells(rng)

    for col_idx, value in enumerate(STRUCTURE_HEADER_ROW, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    for col_idx, value in enumerate(STRUCTURE_LABEL_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = SUBHEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    rows = make_structure_rows()
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGN

        if row[0] is not None:
            fill = VCF_FILL if "MULTI_SAMPLE_GSC" in row[0] else GVCF_FILL
            for col_idx in range(1, 13):
                ws.cell(row=row_idx, column=col_idx).fill = fill
                ws.cell(row=row_idx, column=col_idx).font = HEADER_FONT
        elif row[3] is not None:
            for col_idx in range(1, 13):
                ws.cell(row=row_idx, column=col_idx).fill = SECTION_FILL
                ws.cell(row=row_idx, column=col_idx).font = HEADER_FONT

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42
    for col_idx, width in enumerate(STRUCTURE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main() -> None:
    wb = Workbook()
    write_structure_sheet(wb)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
