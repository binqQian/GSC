from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUT_XLSX = ROOT / "VCF_gVCF字段总表-260424.xlsx"
OUT_MD = ROOT / "VCF_gVCF字段总表-260424.md"

COLUMNS = [
    "序号",
    "格式",
    "层级路径",
    "字段/条目",
    "VCF/gVCF语义字段",
    "类型/编码",
    "必选/条件",
    "写盘顺序/重复",
    "说明",
    "源码依据",
]


def row(
    fmt: str,
    path: str,
    field: str,
    semantic: str,
    encoding: str,
    required: str,
    order: str,
    note: str,
    source: str,
) -> tuple[str, ...]:
    return (fmt, path, field, semantic, encoding, required, order, note, source)


ROWS: list[tuple[str, ...]] = [
    # ---------------------------------------------------------------------
    # Multi-sample VCF/BCF -> .gsc: top-level and main archive.
    # ---------------------------------------------------------------------
    row("VCF/.gsc", "TOP_HEADER", "mode_type", "文件模式", "bool", "Y", "1", "true=lossless，false=lossy；解压 CLI 模式必须匹配。", "src/compressor.cpp::OpenForWriting()/writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "TOP_HEADER", "other_fields_offset", "other-fields 起始偏移", "uint64_t", "Y", "2", "主 archive 结束位置；lossless 时后接 part2 容器。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "TOP_HEADER", "sdsl_offset", "SDSL 尾区起始偏移", "uint64_t", "Y", "3", "从该偏移开始读取 zeros/copies 四个 rrr_vector。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chunk_streams", "chunks_streams_size", "chunk 边界表长度", "uint32_t", "Y", "1", "后续 chunk_stream_item 数量。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chunk_streams/item", "cur_chunk_actual_pos", "累计真实变体数", "uint32_t", "Y", "2.1 * N", "chunk 边界位置，用于 actual variant 定位。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chunk_streams/item", "offset", "chunk 文件偏移", "size_t", "Y", "2.2 * N", "final `.gsc` 里 chunk payload 不带长度前缀，靠该 offset 定位。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "ploidy", "样本倍性", "uint8_t", "Y", "3", "用于 total_haplotypes=n_samples*ploidy。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "max_block_rows", "GT row_block 最大变体数", "uint32_t", "Y", "4", "0 在 reader 侧回退为 total_haplotypes。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "max_block_cols", "GT column_block 最大单倍体数", "uint32_t", "Y", "5", "0 或 >= total_haplotypes 等价于不切列。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "vec_len", "单个 GT 向量字节长", "uint64_t", "Y", "6", "legacy 单列块向量长度；切列时以 col_block 元数据为准。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "no_vec", "GT 向量总数", "uint64_t", "Y", "7", "通常为 2*变体数。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "copy_no / no_copy", "copy 向量数量", "uint64_t", "Y", "8", "压缩侧变量 copy_no，解压侧变量 no_copy。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/copy_origin", "used_bits_cp", "copy-origin 差值 bit 宽", "char", "Y", "9", "读取 copy origin 映射时的位宽。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/copy_origin", "bm_comp_cp_size", "copy-origin payload 字节数", "int32_t writer / int reader", "Y", "10", "后续 bitstream 的 byte size。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/copy_origin", "bm_comp_copy_orgl_id", "copy-origin 差值 bitstream", "bytes by CBitMemory", "Y", "11", "保存 copy 向量到 origin unique id 的差值。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/basic", "n_samples", "样本数", "uint32_t", "Y", "12", "VCF 样本列数量。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/range_index", "chunks_min_pos_size", "chunk 最小 POS 数组长度", "uint32_t", "Y", "13", "控制 chunks_min_pos[] 元素个数。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/range_index", "chunks_min_pos[]", "每个 chunk 的最小 POS", "int64_t[]", "Y", "14 * N", "范围查询粗筛。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chrom_index", "where_chrom_size", "染色体边界项数", "uint32_t", "Y", "15", "后续 where_chrom_item 数量。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chrom_index/item", "chrom_size", "染色体名长度", "size_t", "Y", "16.1 * N", "直接写 ABI size_t。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chrom_index/item", "chrom", "染色体名", "raw char bytes", "Y", "16.2 * N", "不带 NUL，长度由 chrom_size 给出。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chrom_index/item", "chunk_boundary", "染色体边界 chunk id", "int writer / uint32_t reader", "Y", "16.3 * N", "记录 chrom -> chunk boundary。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/column_tiling", "n_col_blocks", "GT 列块数量", "uint32_t", "Y(new) / compat", "17", "旧文件可能没有该段，reader 会回退到 1 个列块。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/column_tiling/item", "start_haplotype", "列块起始单倍体下标", "uint32_t", "Y(new)", "18.1 * N", "第 cb 个列块当前按 cb*max_block_cols 写。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/column_tiling/item", "block_size", "列块单倍体数", "uint32_t", "Y(new)", "18.2 * N", "reader 据此计算 vec_len=(block_size+7)/8。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/permutation", "permutation_count", "排列条目数", "uint32_t", "Y", "19", "新旧排列格式共用 count。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/permutation/item", "row_block_id", "GT row block id", "uint32_t", "Y", "20.1 * N", "新格式和 legacy 都存在。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/permutation/item", "col_block_id", "GT column block id", "uint32_t", "new format only", "20.2 * N", "useLegacyPath=false 时存在。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/permutation/item", "data_size", "排列 payload 长度", "uint32_t", "Y", "20.3 * N", "后续 vint bytes 长度。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/permutation/item", "vint bytes", "单倍体排列数据", "vint_code bytes", "Y", "20.4 * N", "解码后恢复当前 row/col block 的排列。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/meta", "meta_magic", "meta 区魔数", "uint32_t GSC_META_MAGIC", "Y(new) / compat", "21", "新格式用于标记后续 meta_backend；老文件可能直接是 header size。", "src/defs.h; src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/meta", "meta_backend", "meta 压缩后端", "uint8_t backend id", "Y(new)", "22", "header/sample meta 的后端。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/meta", "comp_v_header_size", "压缩 VCF header 字节数", "uint32_t", "Y", "23", "后接 comp_v_header。", "src/compressor.cpp::compress_meta()/writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/meta", "comp_v_header", "VCF header text", "backend-compressed bytes", "Y", "24", "解压后是 append_str(header_text) 产生的 NUL 终止字符串。", "src/compressor.cpp::compress_meta(); src/decompression_reader.cpp::decompress_meta()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/meta", "comp_v_samples_size", "压缩样本列表字节数", "uint32_t", "Y", "25", "后接 comp_v_samples。", "src/compressor.cpp::compress_meta()/writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/meta", "comp_v_samples", "VCF sample names", "backend-compressed bytes", "Y", "26", "解压后是逐个 append_str(sample_name)。", "src/compressor.cpp::compress_meta(); src/decompression_reader.cpp::decompress_meta()"),
    row("VCF/.gsc", "MAIN_ARCHIVE/chunks", "fixed-fields chunk payloads", "CHROM/POS/ID/REF/ALT/QUAL/GT", "concatenated bytes", "Y", "27", "所有 chunk 顺序拼接；每个 chunk 内部见 GSCF fixed-fields chunk。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::setStartChunk()/readFixedFields()"),
    row("VCF/.gsc", "SDSL_TAIL", "zeros_only[0]", "偶数 parity 全零 GT 向量位图", "sdsl::rrr_vector<>", "Y", "1", "从 sdsl_offset 起顺序读取。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "SDSL_TAIL", "zeros_only[1]", "奇数 parity 全零 GT 向量位图", "sdsl::rrr_vector<>", "Y", "2", "GT MSB/LSB 分 parity 管理。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "SDSL_TAIL", "copy_bit_vector[0]", "偶数 parity copy GT 向量位图", "sdsl::rrr_vector<>", "Y", "3", "copy origin 差值在 main archive/copy_origin。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),
    row("VCF/.gsc", "SDSL_TAIL", "copy_bit_vector[1]", "奇数 parity copy GT 向量位图", "sdsl::rrr_vector<>", "Y", "4", "与 zeros_only 一起判断向量解码路径。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReading()"),

    # ---------------------------------------------------------------------
    # Multi-sample fixed-fields chunk and logical VCF fields.
    # ---------------------------------------------------------------------
    row("VCF/.gsc", "GSCF_CHUNK/temp_wrapper", "payload_size", "chunk payload size", "uint64_t", "temp only", "0", "只存在于压缩临时文件；最终 `.gsc` 不保留该长度前缀。", "src/compressor.cpp::compressFixedFieldsChunk(); src/compressor.cpp::writeCompressFlie()"),
    row("VCF/.gsc", "GSCF_CHUNK/header", "magic", "fixed-fields chunk magic", "uint32_t 0x46435347", "Y", "1", "little-endian 观感为 GSCF。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/header", "version", "fixed-fields chunk version", "uint32_t", "Y", "2", "当前 writer 写 GSC_FIXED_FIELDS_RB_VERSION_V2=2。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/header", "total_variants", "chunk 总变体数", "uint32_t", "Y", "3", "chunk 内所有 row_block 的 variant_count 总和。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/header", "row_block_count", "row block 数量", "uint32_t", "Y", "4", "后续目录项数量。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/header", "flags", "预留标记", "uint32_t", "Y", "5", "当前 writer 置 0。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/header_v1", "fixed_fields_gt_off", "chunk-level GT offset", "uint32_t", "V1 compat", "6", "V1 才有；V2 改为每个 row block 自带 gt_off。", "src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/header_v1", "fixed_fields_gt_size", "chunk-level GT size", "uint32_t", "V1 compat", "7", "V1 才有。", "src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "variant_count", "row block 变体数", "uint32_t", "Y", "1 * row_block", "范围查询和 row_block 解码的基本计数。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "first_pos", "row block 首个 POS", "int64_t", "Y", "2 * row_block", "DecoderByRange 依据它剪裁。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::DecoderByRange()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "last_pos", "row block 末个 POS", "int64_t", "Y", "3 * row_block", "DecoderByRange 依据它剪裁。", "src/variant.h; src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::DecoderByRange()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "chrom_off / chrom_size", "VCF CHROM payload 位置/长度", "uint32_t + uint32_t", "Y", "4 * row_block", "指向 data_region 内的 CHROM payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "pos_off / pos_size", "VCF POS payload 位置/长度", "uint32_t + uint32_t", "Y", "5 * row_block", "指向 data_region 内的 POS delta payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "id_off / id_size", "VCF ID payload 位置/长度", "uint32_t + uint32_t", "Y", "6 * row_block", "指向 data_region 内的 ID payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "ref_off / ref_size", "VCF REF payload 位置/长度", "uint32_t + uint32_t", "Y", "7 * row_block", "指向 data_region 内的 REF payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "alt_off / alt_size", "VCF ALT payload 位置/长度", "uint32_t + uint32_t", "Y", "8 * row_block", "指向 data_region 内的 ALT payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "qual_off / qual_size", "VCF QUAL payload 位置/长度", "uint32_t + uint32_t", "Y", "9 * row_block", "指向 data_region 内的 QUAL payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/row_block_dir", "gt_off / gt_size", "VCF FORMAT/GT index payload 位置/长度", "uint32_t + uint32_t", "V2", "10 * row_block", "V2 每个 row block 独立定位 GT payload。", "src/compressor.cpp::writeTempChunkRB(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "CHROM", "VCF fixed column CHROM", "backend-compressed append_str() stream", "Y", "chrom payload", "每条记录一个 NUL 终止字符串。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.cpp::append_str/read_str"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "POS", "VCF fixed column POS", "backend-compressed append(int64_t delta) stream", "Y", "pos payload", "解压后用 read() 读十进制 ASCII+NUL，再与 prev_pos 累加。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.h::append/read"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "ID", "VCF fixed column ID", "backend-compressed append_str() stream", "Y", "id payload", "每条记录一个 NUL 终止字符串。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.cpp::append_str/read_str"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "REF", "VCF fixed column REF", "backend-compressed append_str() stream", "Y", "ref payload", "每条记录一个 NUL 终止字符串。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.cpp::append_str/read_str"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "ALT", "VCF fixed column ALT", "backend-compressed append_str() stream", "Y", "alt payload", "每条记录一个 NUL 终止字符串；复杂多等位支持有限。", "src/block_processing.cpp::addFixedFieldsBlock(); src/compression_reader.cpp::ProcessFixedVariants()"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "QUAL", "VCF fixed column QUAL", "backend-compressed append_str() stream", "Y", "qual payload", "每条记录一个 NUL 终止字符串。", "src/block_processing.cpp::addFixedFieldsBlock(); src/utils.cpp::append_str/read_str"),
    row("VCF/.gsc", "GSCF_CHUNK/data_region", "GT", "VCF FORMAT/GT", "GT index compressed bytes + 1-byte backend marker", "Y", "gt payload", "不是原始 GT 文本；marker 0/1/2 分别表示 bsc/zstd/brotli。", "src/compressor.cpp::compressFixedFieldsChunkToPayload(); src/decompression_reader.cpp::Decoder()"),
    row("VCF/.gsc", "GSCF_CHUNK/legacy", "legacy no_variants", "legacy chunk 变体数", "uint32_t", "compat", "legacy 1", "旧 fixed-fields chunk 不是 GSCF magic 开头。", "src/compressor.cpp::writeTempFlie(); src/decompression_reader.cpp::readFixedFields()"),
    row("VCF/.gsc", "GSCF_CHUNK/legacy", "legacy field entry", "legacy fixed field", "uint32_t size + bytes", "compat", "legacy repeated", "旧格式按 chrom/pos/id/ref/alt/qual/gt 写 size+payload。", "src/compressor.cpp::writeTempFlie(); src/decompression_reader.cpp::readFixedFields()"),

    # ---------------------------------------------------------------------
    # Multi-sample lossless part2 and all VCF other fields.
    # ---------------------------------------------------------------------
    row("VCF/.gsc", "LOSSLESS_PART2/region", "part2_region", "FILTER/INFO/FORMAT 容器", "[other_fields_offset, sdsl_offset)", "lossless only", "after main archive", "由 File_Handle_2 管理；lossy 文件没有 other fields。", "src/compressor.cpp::writeCompressFlie(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/payload", "stream payload area", "各 stream part 本体", "bytes", "lossless only", "1", "所有 stream part 先顺序写 payload，再写 footer。", "src/file_handle.cpp::AddParamsPart()/AddPartComplete()/serialize()"),
    row("VCF/.gsc", "LOSSLESS_PART2/streams", "part2_params", "lossless 元信息流", "stream name + payload", "lossless only", "registered stream", "保存 actual_variants、key 表、backend、FORMAT dictionaries。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/streams", "key_<i>_size", "第 i 个 FILTER/INFO/FORMAT 字段 size 子流", "stream name + compressed uint32_t[]", "lossless only", "for each key", "用于按记录切分字段 data。", "src/compression_reader.cpp::InitVarinats(); src/compressor.cpp::processOtherFieldsResult()"),
    row("VCF/.gsc", "LOSSLESS_PART2/streams", "key_<i>_data", "第 i 个 FILTER/INFO/FORMAT 字段 data 子流", "stream name + compressed bytes", "lossless only", "for each key", "保存字段序列化值；特殊 FORMAT 字段有额外 codec。", "src/compression_reader.cpp::InitVarinats(); src/compressor.cpp::processOtherFieldsResult()"),
    row("VCF/.gsc", "LOSSLESS_PART2/footer", "n_streams", "stream 数量", "File_Handle_2::Write(size_t)", "lossless only", "footer 1", "Write(size_t)=[no_bytes:u8][big-endian bytes]。", "src/file_handle.cpp::serialize()/Write(size_t)"),
    row("VCF/.gsc", "LOSSLESS_PART2/footer", "stream_name", "stream 名", "NUL-terminated string", "lossless only", "footer 2 * stream", "例如 part2_params、key_0_size、key_0_data。", "src/file_handle.cpp::serialize()/Write(string)"),
    row("VCF/.gsc", "LOSSLESS_PART2/footer", "part_count", "stream part 数", "File_Handle_2::Write(size_t)", "lossless only", "footer 3 * stream", "该 stream 由多少个 part 组成。", "src/file_handle.cpp::serialize()/deserialize()"),
    row("VCF/.gsc", "LOSSLESS_PART2/footer", "part.offset", "part payload 偏移", "File_Handle_2::Write(size_t)", "lossless only", "footer 4 * part", "相对 part2 payload 起点，不是整个 `.gsc` 文件偏移。", "src/file_handle.cpp::serialize()/deserialize()"),
    row("VCF/.gsc", "LOSSLESS_PART2/footer", "part.size", "part payload 长度", "File_Handle_2::Write(size_t)", "lossless only", "footer 5 * part", "用于 GetPart 定位读取。", "src/file_handle.cpp::serialize()/deserialize()"),
    row("VCF/.gsc", "LOSSLESS_PART2/footer", "footer_size", "footer 字节数", "fixed 8 raw bytes", "lossless only", "EOF of part2", "WriteFixed(size_t) 固定写 8 字节，依赖 64-bit little-endian ABI。", "src/file_handle.cpp::WriteFixed()/read_fixed()"),
    row("VCF/.gsc", "LOSSLESS_PART2/part2_params", "backend_id prefix", "part2_params payload 压缩后端", "uint8_t", "lossless only", "payload 1", "part2_params 第 1 字节，后面是压缩后的 v_desc。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/part2_params", "compressed v_desc", "part2_params 主体", "backend-compressed bytes", "lossless only", "payload 2", "解压后得到 append()/append_str() token 流。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "actual_variants_size", "actual_variants 数组长度", "append(uint32_t)", "lossless only", "1", "append() 实际是十进制 ASCII + NUL。", "src/compressor.cpp::CompressProcess(); src/utils.h::append/read"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "actual_variants[]", "每个 chunk 的真实变体数", "append(uint32_t)[]", "lossless only", "2 * N", "多等位拆分时与原始 VCF 记录数不一定 1:1。", "src/compressor.cpp::CompressProcess(); src/compression_reader.cpp::GetActualVariants()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "no_keys", "FILTER/INFO/FORMAT key 总数", "append(uint32_t)", "lossless only", "3", "后续 key entry 数量。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "key_gt_id", "GT key id", "append(int)", "lossless only", "4", "GT 在 key 表中的位置。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/key_entry", "key_id", "字段内部 id", "append(uint32_t)", "lossless only", "5.1 * key", "key_desc.key_id。", "src/variant.h::key_desc; src/compressor.cpp::CompressProcess()"),
    row("VCF/.gsc", "LOSSLESS_PART2/key_entry", "actual_field_id", "拓扑排序后的字段 id", "append(uint32_t)", "lossless only", "5.2 * key", "由 field_order_graph/topo_sort 修正字段顺序。", "src/variant.h::key_desc; src/compression_reader.cpp::UpdateKeys()"),
    row("VCF/.gsc", "LOSSLESS_PART2/key_entry", "keys_type", "字段类别", "append(uint64_t enum key_type_t)", "lossless only", "5.3 * key", "flt=FILTER，info=INFO，fmt=FORMAT。", "src/variant.h::key_type_t/key_desc; src/compressor.cpp::CompressProcess()"),
    row("VCF/.gsc", "LOSSLESS_PART2/key_entry", "type", "BCF 字段值类型", "append(int8_t)", "lossless only", "5.4 * key", "BCF_HT_FLAG/INT/REAL/STR。", "src/variant.h::key_desc; src/compressor.cpp::CompressProcess()"),
    row("VCF/.gsc", "LOSSLESS_PART2/key_entry", "name", "字段名", "append_str()", "lossless only", "5.5 * key", "如 DP、AD、PL、GQ、PGT、PID 或任意 header key。", "src/variant.h::key_desc; src/compressor.cpp::CompressProcess()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "params.backend", "part2 内记录的压缩后端", "append(uint32_t)", "lossless only", "6", "和 backend_id prefix 是两层信息。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "fmt_dict_version", "FORMAT 字典版本", "append(uint32_t)", "lossless only", "7", "当前 writer 固定写 1；旧文件可缺失。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "fmt_dict_size", "FORMAT 字典 blob 大小", "append(uint32_t)", "lossless only", "8", "后接 fmt_dict_blob。", "src/compressor.cpp::CompressProcess(); src/decompression_reader.cpp::OpenReadingPart2()"),
    row("VCF/.gsc", "LOSSLESS_PART2/v_desc", "fmt_dict_blob", "AD/PL/PID 字典", "bytes", "lossless only", "9", "FmtDictionaries 序列化结果，用于特殊 FORMAT codec 复原。", "src/fmt_compress/fmt_dictionaries.h; src/compressor.cpp::CompressProcess()"),
    row("VCF/.gsc", "VCF_LOGICAL_FIELDS", "FILTER", "VCF FILTER 字段全集", "part2 key type=flt", "lossless only", "per record/key", "lossless 保存 header 中出现的 FILTER key；lossy 不保留。", "src/compression_reader.cpp::InitVarinats()/GetVariantFromRec(); src/variant.h::key_desc"),
    row("VCF/.gsc", "VCF_LOGICAL_FIELDS", "INFO", "VCF INFO 字段全集", "part2 key type=info", "lossless only", "per record/key", "lossless 保存 header 中出现的 INFO key；FLAG/INT/REAL/STR 按 type 解码。", "src/compression_reader.cpp::InitVarinats()/GetVariantFromRec(); src/variant.h::key_desc"),
    row("VCF/.gsc", "VCF_LOGICAL_FIELDS", "FORMAT", "VCF FORMAT 字段全集", "part2 key type=fmt", "lossless only", "per record/key", "GT 单独走 GT bitplane/index；其他 FORMAT 走 part2，部分字段有特殊 codec。", "src/compression_reader.cpp::InitVarinats()/GetVariantFromRec(); src/variant.h::key_desc"),
    row("VCF/.gsc", "VCF_INTERNAL/variant_desc_t", "alt_num", "ALT 等位数量", "int", "internal", "not directly serialized", "用于多等位处理；不是独立 on-disk 字段。", "src/variant.h::variant_desc_t; src/compression_reader.cpp::ProcessFixedVariants()"),
    row("VCF/.gsc", "VCF_INTERNAL/field_desc", "present", "字段是否存在", "bool", "internal/lossless", "per record/key", "当前记录是否有该 FILTER/INFO/FORMAT 字段。", "src/variant.h::field_desc; src/compression_reader.cpp::GetVariantFromRec()"),
    row("VCF/.gsc", "VCF_INTERNAL/field_desc", "data_size", "字段 payload 长度", "uint32_t", "internal/lossless", "per record/key", "写入 key_i_size 子流。", "src/variant.h::field_desc; src/compressor.cpp::processOtherFieldsResult()"),
    row("VCF/.gsc", "VCF_INTERNAL/field_desc", "data", "字段 payload", "char*", "internal/lossless", "per record/key", "写入 key_i_data 子流。", "src/variant.h::field_desc; src/compressor.cpp::processOtherFieldsResult()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "GT", "FORMAT/GT", "2 bitplanes per haplotype + sparse index", "Y", "fixed-fields GT", "MSB/LSB: 00=0, 01=1, 10=., 11=2；每个变体贡献两条向量。", "src/compression_reader.cpp::addVariant(); src/decompressor.cpp::initialLut()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "AD", "FORMAT/AD", "special stream codec 0/5/7", "lossless if present", "part2 key_i_data", "AD 有字典/模式识别优化；codec 7 是当前优先路径。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "DP", "FORMAT/DP", "record codec 0/1", "lossless if present", "part2 key_i_data", "可由 AD sum 预测，保存 raw/exceptions。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "MIN_DP", "FORMAT/MIN_DP", "record codec 0/1", "lossless if present", "part2 key_i_data", "可由 DP 预测，保存 exceptions 或 raw fallback。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "PL", "FORMAT/PL", "special stream codec 0/6", "lossless if present", "part2 key_i_data", "识别常见 PL pattern，否则使用字典/归一化复原。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "GQ", "FORMAT/GQ", "record codec 0/1", "lossless if present", "part2 key_i_data", "可由 PL 二小值/特定 pattern 预测。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "PGT", "FORMAT/PGT", "special stream codec 0/1", "lossless if present", "part2 key_i_data", "稀疏存储非 `.` 短字符串。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "VCF_FORMAT_SPECIAL", "PID", "FORMAT/PID", "special stream codec 0/1/2", "lossless if present", "part2 key_i_data", "稀疏 + 字典，依赖 fmt_dict_blob。", "src/compressor.cpp::processOtherFieldsResult(); src/decompression_reader.cpp::decompress_other_fileds()"),
    row("VCF/.gsc", "GT_SPARSE_INDEX", "delta list + 0 terminator", "GT unique vector 稀疏行", "vint_code::EncodeArray()", "Y", "inside GT payload", "每条非 zero/copy 向量保存 1 的位置 delta，行尾 0 终止。", "src/block_processing.cpp; src/decompressor.cpp::decoded_vector_row(); src/vint_code.cpp"),
    row("VCF/.gsc", "GT_SPARSE_INDEX", "byte-local XOR domain", "GT 差分域", "byte 内相邻 bit XOR", "Y", "before sparse index", "解压由 initialXORLut() 逆变换。", "src/decompressor.cpp::initialXORLut(); src/block_processing.cpp"),

    # ---------------------------------------------------------------------
    # Native single-sample gVCF file envelope.
    # ---------------------------------------------------------------------
    row("gVCF/native", "FILE_HEADER", "magic", "文件魔数", "uint32_t 0x47564346", "Y", "1", "GVCF_FILE_MAGIC，ASCII 为 GVCF。", "src/gvcf/gvcf_compressor.h; src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER", "version", "文件版本", "uint32_t", "Y", "2", "当前 writer 写 GVCF_FILE_VERSION=4；reader 支持 <=4。", "src/gvcf/gvcf_compressor.h; src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER", "backend_id", "压缩后端", "uint8_t", "V2+", "3", "V1 默认 zstd；V2+ 显式保存 backend。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER", "num_samples", "样本数", "uint32_t", "Y", "4", "gVCF 主线通常单样本，但格式保留计数。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/header_text", "header_flag", "header 是否压缩", "uint8_t", "V3+", "5", "1=压缩 header；0=raw header。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/header_text", "original_size", "header 原始长度", "uint32_t", "if header_flag=1", "6.1", "用于解压 buffer 预期大小。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/header_text", "compressed_size", "header 压缩长度", "uint32_t", "if header_flag=1", "6.2", "后接 compressed_header bytes。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/header_text", "compressed_header", "VCF/gVCF header text", "backend-compressed bytes", "if header_flag=1", "6.3", "当压缩后更小时写该分支。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/header_text", "header_size", "raw header 长度", "uint32_t", "if header_flag=0 or V1/V2", "6.1 alt", "后接 raw header bytes。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/header_text", "raw_header", "VCF/gVCF header text", "raw bytes", "if header_flag=0 or V1/V2", "6.2 alt", "未压缩 header 文本。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/sample_names", "name_size", "样本名长度", "uint32_t", "Y", "7.1 * num_samples", "后接 sample name bytes。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER/sample_names", "name", "样本名", "raw bytes", "Y", "7.2 * num_samples", "不带 NUL，长度由 name_size 给出。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "FILE_HEADER", "placeholder", "头部尾占位", "uint64_t", "Y", "8", "reader 直接跳过；当前不承载逻辑。", "src/gvcf/gvcf_compressor.cpp::WriteFileHeader()/ReadFileHeader()"),
    row("gVCF/native", "BLOCK_AREA/block_entry", "block_size", "单块 payload 长度", "uint32_t", "Y", "1 * block", "后接 CompressedGVCFBlock 序列化结果。", "src/gvcf/gvcf_compressor.cpp::CompressAndWriteBlock()/ReadBlock()"),
    row("gVCF/native", "BLOCK_AREA/block_entry", "block_payload", "CompressedGVCFBlock", "bytes", "Y", "2 * block", "块内再带 GVCF magic、version、字段 wrapper。", "src/gvcf/gvcf_compressor.cpp::CompressAndWriteBlock(); src/gvcf/gvcf_block.cpp::Serialize()"),
    row("gVCF/native", "FILE_FOOTER", "num_blocks", "块数量", "uint32_t", "Y", "1", "footer 起始字段。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),
    row("gVCF/native", "FILE_FOOTER", "block_offsets[]", "块偏移表", "uint64_t[]", "Y", "2 * num_blocks", "每项指向对应 block_size 字段的位置。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),
    row("gVCF/native", "FILE_FOOTER", "total_variants", "总变体数", "uint64_t", "Y", "3", "全文件 variant 统计。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),
    row("gVCF/native", "FILE_FOOTER/block_index", "num_indices", "范围索引条目数", "uint32_t", "V4", "4", "V4 新增，用于 gvcf-query。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader(); src/gvcf/gvcf_compressor.cpp::GVCFQueryer::Open()"),
    row("gVCF/native", "FILE_FOOTER/block_index/item", "chrom_len", "索引染色体名长度", "uint32_t", "V4", "5.1 * index", "后接 chrom bytes。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),
    row("gVCF/native", "FILE_FOOTER/block_index/item", "chrom", "索引染色体名", "raw bytes", "V4", "5.2 * index", "块覆盖的染色体。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),
    row("gVCF/native", "FILE_FOOTER/block_index/item", "start_pos", "块起始 POS", "uint64_t", "V4", "5.3 * index", "范围查询筛选字段。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),
    row("gVCF/native", "FILE_FOOTER/block_index/item", "end_pos", "块结束 POS/END", "uint64_t", "V4", "5.4 * index", "优先用最后一条 INFO/END，否则用 POS。", "src/gvcf/gvcf_compressor.cpp::CompressAndWriteBlock()/WriteFileFooter()"),
    row("gVCF/native", "FILE_FOOTER/block_index/item", "file_offset", "块文件偏移", "uint64_t", "V4", "5.5 * index", "指向 block_size。", "src/gvcf/gvcf_compressor.cpp::CompressAndWriteBlock()/WriteFileFooter()"),
    row("gVCF/native", "FILE_FOOTER/block_index/item", "variant_count", "块内变体数", "uint32_t", "V4", "5.6 * index", "用于查询输出裁剪。", "src/gvcf/gvcf_compressor.h::BlockIndex; src/gvcf/gvcf_compressor.cpp::WriteFileFooter()"),
    row("gVCF/native", "FILE_FOOTER", "footer_offset", "footer 起始偏移", "uint64_t at EOF", "Y", "EOF", "文件最后 8 字节；reader 先回读它再 seek 到 footer。", "src/gvcf/gvcf_compressor.cpp::WriteFileFooter()/ReadFileHeader()"),

    # ---------------------------------------------------------------------
    # gVCF logical and compressed block fields.
    # ---------------------------------------------------------------------
    row("gVCF/native", "GVCFBlock/raw_position", "position.chrom", "CHROM", "vector<string>", "Y", "raw before compression", "gVCF block 原始字段。", "src/gvcf/gvcf_block.h::GVCFPositionFields"),
    row("gVCF/native", "GVCFBlock/raw_position", "position.pos", "POS", "vector<uint64_t>", "Y", "raw before compression", "gVCF block 原始字段。", "src/gvcf/gvcf_block.h::GVCFPositionFields"),
    row("gVCF/native", "GVCFBlock/raw_position", "position.id", "ID", "vector<string>", "Y", "raw before compression", "通常为 `.`。", "src/gvcf/gvcf_block.h::GVCFPositionFields"),
    row("gVCF/native", "GVCFBlock/raw_sequence", "sequence.ref", "REF", "vector<string>", "Y", "raw before compression", "参考碱基/序列。", "src/gvcf/gvcf_block.h::GVCFSequenceFields"),
    row("gVCF/native", "GVCFBlock/raw_sequence", "sequence.alt", "ALT", "vector<vector<string>>", "Y", "raw before compression", "支持 multi-allelic ALT；压缩时主路径以第一 ALT 为主，额外 ALT 另存。", "src/gvcf/gvcf_block.h::GVCFSequenceFields; src/gvcf/gvcf_field_compress.cpp::AltCompressor::Compress()"),
    row("gVCF/native", "GVCFBlock/raw_quality", "quality.qual", "QUAL", "vector<float>", "Y", "raw before compression", "质量值。", "src/gvcf/gvcf_block.h::GVCFQualityFields"),
    row("gVCF/native", "GVCFBlock/raw_quality", "quality.filter", "FILTER", "vector<string>", "Y", "raw before compression", "过滤字段。", "src/gvcf/gvcf_block.h::GVCFQualityFields"),
    row("gVCF/native", "GVCFBlock/raw_info", "info.end", "INFO/END", "vector<int64_t>", "if present", "raw before compression", "gVCF block end position。", "src/gvcf/gvcf_block.h::GVCFInfoFields"),
    row("gVCF/native", "GVCFBlock/raw_sample", "sample.gt", "FORMAT/GT", "vector<GenotypeData>", "Y", "raw before compression", "GenotypeData={allele1, allele2, phased}。", "src/gvcf/gvcf_block.h::GenotypeData/GVCFSampleFields"),
    row("gVCF/native", "GVCFBlock/raw_sample", "sample.dp", "FORMAT/DP", "vector<int32_t>", "if present", "raw before compression", "read depth。", "src/gvcf/gvcf_block.h::GVCFSampleFields"),
    row("gVCF/native", "GVCFBlock/raw_sample", "sample.gq", "FORMAT/GQ", "vector<int32_t>", "if present", "raw before compression", "genotype quality。", "src/gvcf/gvcf_block.h::GVCFSampleFields"),
    row("gVCF/native", "GVCFBlock/raw_sample", "sample.min_dp", "FORMAT/MIN_DP", "vector<int32_t>", "if present", "raw before compression", "minimum depth in gVCF block。", "src/gvcf/gvcf_block.h::GVCFSampleFields"),
    row("gVCF/native", "GVCFBlock/raw_sample", "sample.pl", "FORMAT/PL", "vector<vector<int32_t>>", "if present", "raw before compression", "phred-scaled likelihoods。", "src/gvcf/gvcf_block.h::GVCFSampleFields"),
    row("gVCF/native", "GVCFBlock/raw_sample", "sample.ad", "FORMAT/AD", "vector<vector<int32_t>>", "if present", "raw before compression", "allelic depths。", "src/gvcf/gvcf_block.h::GVCFSampleFields"),
    row("gVCF/native", "GVCFBlock/raw_unknown", "unknown_info[name]", "任意其他 INFO 字段", "unordered_map<string, vector<string>>", "if present", "raw before compression", "generic dictionary codec。", "src/gvcf/gvcf_block.h::GVCFBlock; src/gvcf/gvcf_field_compress.cpp::GenericFieldCompressor"),
    row("gVCF/native", "GVCFBlock/raw_unknown", "unknown_format[name]", "任意其他 FORMAT 字段", "unordered_map<string, vector<string>>", "if present", "raw before compression", "generic dictionary codec。", "src/gvcf/gvcf_block.h::GVCFBlock; src/gvcf/gvcf_field_compress.cpp::GenericFieldCompressor"),
    row("gVCF/native", "CompressedGVCFBlock/header", "block magic", "块内魔数", "4 bytes: G V C F", "Y", "1", "每个 block payload 内部自带二级 magic。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/header", "block version", "块版本", "uint8_t", "Y", "2", "当前固定写 1；反序列化只接受 1。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/header", "variant_count", "块内变体数", "VarUint", "Y", "3", "压缩块元数据。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/header", "sample_count", "块内样本数", "VarUint", "Y", "4", "压缩块元数据。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/header", "flags", "字段存在标记", "uint8_t", "Y", "5", "bit0=has_end_field，bit1=has_min_dp。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "chrom", "CHROM", "CompressedField", "Y", "6", "由 ChromCompressor 压缩，方法 RLE。", "src/gvcf/gvcf_block.h::CompressedGVCFBlock; src/gvcf/gvcf_field_compress.cpp"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "pos", "POS", "CompressedField", "Y", "7", "由 PosCompressor 压缩，方法 DELTA。", "src/gvcf/gvcf_block.h::CompressedGVCFBlock; src/gvcf/gvcf_field_compress.cpp"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "id", "ID", "CompressedField", "Y", "8", "MASK(`.` dominant) 或 DICTIONARY。", "src/gvcf/gvcf_field_compress.cpp::IdCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "ref", "REF", "CompressedField", "Y", "9", "单碱基 2-bit + exceptions。", "src/gvcf/gvcf_field_compress.cpp::RefCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "alt", "ALT", "CompressedField", "Y", "10", "MASK 第一 ALT + extra_alts 扩展。", "src/gvcf/gvcf_field_compress.cpp::AltCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "qual", "QUAL", "CompressedField", "Y", "11", "raw float bytes + optional backend wrapper。", "src/gvcf/gvcf_field_compress.cpp::QualCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "filter", "FILTER", "CompressedField", "Y", "12", "RLE / MASK / DICTIONARY 自适应。", "src/gvcf/gvcf_field_compress.cpp::FilterCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/known_fields", "info_end", "INFO/END", "CompressedField", "flags & 0x01", "13", "END 可由相邻 POS 推断时只存 exceptions。", "src/gvcf/gvcf_field_compress.cpp::EndCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "gt_mask", "FORMAT/GT dominant mask", "CompressedField", "Y", "14", "GT string MASK，dominant 默认 0/0。", "src/gvcf/gvcf_field_compress.cpp::GTCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "gt_patches", "FORMAT/GT patches", "CompressedField", "Y", "15", "当前序列化保留该字段；GT patch 信息主要已经在 gt_mask 的 MaskResult 内。", "src/gvcf/gvcf_block.cpp::SerializeField(); src/gvcf/gvcf_field_compress.cpp::GTCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "gt_phase", "FORMAT/GT phase", "CompressedField", "Y", "16", "相位 bool bitmap 经 RLEByte 编码。", "src/gvcf/gvcf_field_compress.cpp::GTCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "dp", "FORMAT/DP", "CompressedField", "Y/if present", "17", "若 MIN_DP 存在，MinDPCompressor 同时写 dp 和 dp_min_dp_diff。", "src/gvcf/gvcf_field_compress.cpp::DPCompressor/MinDPCompressor"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "gq", "FORMAT/GQ", "CompressedField", "if present", "18", "由 PL 预测并保存 exceptions。", "src/gvcf/gvcf_field_compress.cpp::GQCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "min_dp", "FORMAT/MIN_DP standalone", "CompressedField", "flags & 0x02", "19", "Serialize 顺序中存在；当前压缩主路径主要写 dp_min_dp_diff + dp。", "src/gvcf/gvcf_block.cpp::Serialize(); src/gvcf/gvcf_field_compress.cpp::MinDPCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "dp_min_dp_diff", "DP-MIN_DP", "CompressedField", "flags & 0x02", "20", "差值多为 0 时用 MASK，否则 RLE。", "src/gvcf/gvcf_field_compress.cpp::MinDPCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "pl", "FORMAT/PL", "CompressedField", "if present", "21", "PL optimized v2：标准三元组 bitmask + PL1/PL2 + exception dict。", "src/gvcf/gvcf_field_compress.cpp::PLCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/sample_fields", "ad", "FORMAT/AD", "CompressedField", "if present", "22", "当前复用 PLCompressor 的矢量整数模式。", "src/gvcf/gvcf_field_compress.cpp::ADCompressor::Compress()"),
    row("gVCF/native", "CompressedGVCFBlock/unknown_info", "unknown_info_count", "未知 INFO 字段数量", "VarUint", "Y", "23", "后续重复 name + CompressedField。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/unknown_info/item", "name_len / name / field", "未知 INFO 字段项", "VarUint + bytes + CompressedField", "if count>0", "24 * N", "name 是 INFO key。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/unknown_format", "unknown_format_count", "未知 FORMAT 字段数量", "VarUint", "Y", "25", "后续重复 name + CompressedField。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedGVCFBlock/unknown_format/item", "name_len / name / field", "未知 FORMAT 字段项", "VarUint + bytes + CompressedField", "if count>0", "26 * N", "name 是 FORMAT key。", "src/gvcf/gvcf_block.cpp::CompressedGVCFBlock::Serialize()/Deserialize()"),
    row("gVCF/native", "CompressedField", "method", "字段压缩方法", "uint8_t FieldCompressionMethod", "Y", "1", "NONE=0,RLE=1,DELTA=2,MASK=3,DICTIONARY=4,DELTA_RLE=5,MASK_RLE=6,ADAPTIVE=7。", "src/gvcf/gvcf_block.h::FieldCompressionMethod; src/gvcf/gvcf_block.cpp::SerializeField()"),
    row("gVCF/native", "CompressedField", "original_count", "原始元素个数", "VarUint", "Y", "2", "用于字段解压后校验/重建。", "src/gvcf/gvcf_block.cpp::SerializeField()/DeserializeField()"),
    row("gVCF/native", "CompressedField", "data_size", "字段 data 长度", "VarUint", "Y", "3", "后接 data。", "src/gvcf/gvcf_block.cpp::SerializeField()/DeserializeField()"),
    row("gVCF/native", "CompressedField", "data", "字段编码 payload", "bytes", "Y", "4", "FieldCompressor 还会在 data 内前置 1 字节 backend flag：1=后端压缩，0=未后端压缩。", "src/gvcf/gvcf_block.cpp::SerializeField(); src/gvcf/gvcf_field_compress.cpp::FieldCompressor::ApplyBackendCompression()"),

    # ---------------------------------------------------------------------
    # gVCF lower-level encoding payloads.
    # ---------------------------------------------------------------------
    row("gVCF/native", "ENCODING/varint", "VarUint", "无符号变长整数", "7-bit continuation varint", "as used", "-", "低 7 bit 存值，高位 bit 为 continuation。", "src/gvcf/gvcf_encoding.cpp::VarIntUtil"),
    row("gVCF/native", "ENCODING/varint", "VarInt", "有符号变长整数", "ZigZag + VarUint", "as used", "-", "有符号差值先 zigzag。", "src/gvcf/gvcf_encoding.cpp::VarIntUtil"),
    row("gVCF/native", "ENCODING/RLE_string", "RLE(string)", "字符串 RLE", "[original_len][run_count]{[strlen][str bytes][count]}", "as used", "-", "CHROM 等字符串字段可用。", "src/gvcf/gvcf_encoding.cpp::RLEEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/RLE_int", "RLE(int)", "整数 RLE", "[original_len][run_count]{[value_varint][count]}", "as used", "-", "DP、indices 等可用。", "src/gvcf/gvcf_encoding.cpp::RLEIntEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/RLE_byte", "RLE(byte)", "字节 RLE", "[original_len][run_count]{[value:u8][count]}", "as used", "-", "bitmask/phase bitmap 常用。", "src/gvcf/gvcf_encoding.cpp::RLEByteEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/Delta", "Delta", "整数 delta", "[original_len][first_value][delta_count][delta_varint...]", "as used", "-", "POS、索引等常用。", "src/gvcf/gvcf_encoding.cpp::DeltaEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/Mask_string", "Mask(string)", "字符串 dominant+patch", "[original_len][dominant_len][dominant][mask_size][RLEByte(mask)][patch_count][idx_size?][Delta(indices)][patch strings...]", "as used", "-", "ID、ALT、GT string、FILTER 等可用。", "src/gvcf/gvcf_encoding.cpp::MaskEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/Mask_int", "Mask(int)", "整数 dominant+patch", "[original_len][dominant_value][mask_size][RLEByte(mask)][patch_count][idx_size?][Delta(indices)][patch varint...]", "as used", "-", "DP-MIN_DP diff 等可用。", "src/gvcf/gvcf_encoding.cpp::MaskIntEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/Dict_string", "Dictionary(string)", "字符串字典", "[original_len][dict_size]{[len][bytes]}[use_rle][rle_size+RLEInt(indices) | idx_count+idx...]", "as used", "-", "unknown INFO/FORMAT、ID/FILTER 等可用。", "src/gvcf/gvcf_encoding.cpp::DictEncoder::Serialize()"),
    row("gVCF/native", "ENCODING/Dict_int", "Dictionary(int)", "整数字典", "[original_len][dict_size]{[value_varint]}[use_rle][rle_size+RLEInt(indices) | idx_count+idx...]", "as used", "-", "DP 等整数离散字段可用。", "src/gvcf/gvcf_encoding.cpp::DictIntEncoder::Serialize()"),
    row("gVCF/native", "FIELD_PAYLOAD/REF", "REF 2-bit payload", "REF field specialized payload", "[exception_count][delta_indices][exception strings][packed_bases_size][packed_bases]", "Y for REF", "field data before backend flag", "A/C/G/T 单碱基 2-bit 打包，非单碱基作为 exception。", "src/gvcf/gvcf_field_compress.cpp::RefCompressor::Compress()"),
    row("gVCF/native", "FIELD_PAYLOAD/ALT", "ALT extra_alts payload", "ALT multi-allelic supplement", "Mask(first_alt) + [has_extra:u8][extra_count][record_idx][extra_count][len+bytes...]", "Y for ALT", "field data before backend flag", "第一 ALT 走 Mask，额外 ALT 追加保存。", "src/gvcf/gvcf_field_compress.cpp::AltCompressor::Compress()"),
    row("gVCF/native", "FIELD_PAYLOAD/QUAL", "QUAL raw floats", "QUAL field payload", "[count][float32 little-endian bytes...]", "Y for QUAL", "field data before backend flag", "float 直接拆成 4 个 byte。", "src/gvcf/gvcf_field_compress.cpp::QualCompressor::Compress()"),
    row("gVCF/native", "FIELD_PAYLOAD/END", "END inference exceptions", "INFO/END payload", "[exception_count][delta_exception_indices...][end_minus_pos values...]", "if END present", "field data before backend flag", "可由下一条 POS 推断的 END 不写具体值。", "src/gvcf/gvcf_field_compress.cpp::EndCompressor::Compress()"),
    row("gVCF/native", "FIELD_PAYLOAD/GT", "GT phase payload", "GT phase field payload", "[count][RLEByte(phase_bitmap)]", "Y", "gt_phase data before backend flag", "phased=true 的位置在 bitmap 中置位。", "src/gvcf/gvcf_field_compress.cpp::GTCompressor::Compress()"),
    row("gVCF/native", "FIELD_PAYLOAD/GQ", "GQ prediction exceptions", "FORMAT/GQ payload", "[count][bitmask_size][RLEByte(predicted_mask)][exception_count][idx_size?][Delta(indices)][exception values...]", "if GQ present", "field data before backend flag", "预测值来自 PL；不匹配处保存 exception。", "src/gvcf/gvcf_field_compress.cpp::GQCompressor::Compress()"),
    row("gVCF/native", "FIELD_PAYLOAD/PL_AD", "PL/AD optimized v2", "FORMAT/PL and FORMAT/AD payload", "[version=2][count][standard_bitmask][standard_count][PL1...][PL2...][exception_dict][exception_indices]", "if PL/AD present", "field data before backend flag", "AD 当前复用 PLCompressor，因此 payload 形状相同。", "src/gvcf/gvcf_field_compress.cpp::PLCompressor::Compress(); ADCompressor::Compress()"),
]


def md_escape(value: object) -> str:
    text = str(value)
    return text.replace("|", "\\|").replace("\n", "<br>")


def write_markdown(rows: list[tuple[object, ...]]) -> None:
    lines: list[str] = [
        "# VCF/gVCF 字段总表（260424）",
        "",
        "本表参考 `docs/0_avsStructure/` 的二进制封装表格风格，但字段顺序、类型、条件和语义以当前源码实际 writer/reader 为准。",
        "",
        "覆盖范围：",
        "",
        "- 多样本 VCF/BCF 压缩输出的 `.gsc`：顶层头、主 archive、GSCF fixed-fields chunk、lossless part2、GT zero/copy 索引。",
        "- 单样本 `gsc gvcf` 输出的 native `GVCF`：文件头、块区、footer、CompressedGVCFBlock、CompressedField、基础编码器。",
        "",
        "| " + " | ".join(COLUMNS) + " |",
        "| " + " | ".join("---" for _ in COLUMNS) + " |",
    ]
    for row_values in rows:
        lines.append("| " + " | ".join(md_escape(v) for v in row_values) + " |")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_workbook(rows: list[tuple[object, ...]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "字段总表"
    ws.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", horizontal="left", wrap_text=True)

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row_idx, values in enumerate(rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    widths = [8, 14, 34, 28, 28, 42, 16, 18, 64, 58]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    info = wb.create_sheet("说明")
    info_rows = [
        ("文档定位", "当前项目 VCF/.gsc 与 native gVCF on-disk 字段总表。"),
        ("参考材料", "参考 docs/0_avsStructure 的表格风格；不把 AVS-G V1 目标格式当成当前事实格式。"),
        ("源码基准", "多样本以 src/compressor.cpp、src/decompression_reader.cpp、src/file_handle.cpp、src/variant.h 为准；gVCF 以 src/gvcf/gvcf_compressor.cpp、src/gvcf/gvcf_block.cpp、src/gvcf/gvcf_field_compress.cpp、src/gvcf/gvcf_encoding.cpp 为准。"),
        ("当前版本", "多样本 fixed-fields chunk 当前 writer 为 GSCF V2；native gVCF 当前 writer 为 GVCF_FILE_VERSION=4。"),
        ("注意", ".gsc 主 archive 含 bool、size_t 和固定 8 字节 size_t footer_size，默认依赖 64-bit little-endian ABI。"),
    ]
    for col_idx, name in enumerate(["对象", "说明"], start=1):
        cell = info.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row_idx, values in enumerate(info_rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = info.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 100
    info.freeze_panes = "A2"

    wb.save(OUT_XLSX)


def main() -> None:
    numbered_rows = [(idx, *values) for idx, values in enumerate(ROWS, start=1)]
    write_markdown(numbered_rows)
    write_workbook(numbered_rows)
    print(f"Wrote {OUT_MD}")
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
